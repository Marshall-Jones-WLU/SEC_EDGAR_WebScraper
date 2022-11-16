'''
Author: Marshall Jones
Filename: 10kAnalysis

Description:
The user inputs the ticker symbol of a company. It then finds that
company's SEC Edgar CIK number. It then finds the Accession number of the company's most recently
published 10-k filing (the type of filing and number of most recently released). From
there, the get_filing_summary() function returns the summary of the filing in xml format. The
parse_filing_summary() function parses the filing summary and creates a list of all the different
reports, or sections of the filing. The grab_financial_statements() function finds the financial
statements that we want from the list of sections of the filing, and then returns a smaller list
of just the financial statement data that we want (currently the Balance Sheet, Income Statement,
Statement of Cash Flows, and Statement of Stockholders' Equity, but more can be added).

The scrape_financial_statements() function takes the data from each financial statement and organizes
it into a structure with column headers, section headers (on each row), and their associated data.
Then, there is another function for each statement that uses the pandas library to put the data into
a data frame, which can be viewed in an excel file. A new worksheet is saved for each of these
financial statements. These functions are called make_balSheet_df(), make_income_df(), make_cashFlow_df(),
and make_equity_df()

TODO:
- Simplify program with classes
- 
'''
# import libraries
import os
import csv
import re, requests
import pandas as pd
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

"""
Use input ticker to find the CIK number, then the Accession number for the most
recent 10k report, as well as the company's info
"""
def getCIK(ticker):
    '''This function uses requests to go to the URL address, and then uses the input ticker to find the
    CIK of that company. It creates a list of CIK's in the variable 'results' but the elements are all the
    same.'''
    URL = 'https://www.sec.gov/cgi-bin/browse-edgar?CIK={}&Find=Search&owner=exclude&action=getcompany'
    CIK_RE = re.compile(r'.*CIK=(\d{10}).*')
    cik_dict = {}
    f = requests.get(URL.format(ticker), stream = True)
    results = CIK_RE.findall(f.text)
    
    if len(results):
        results[0] = int(re.sub('\.[0]*', '.', results[0]))
        cik_dict[str(ticker).upper()] = str(results[0])
        
    CIK = cik_dict[ticker]
    return CIK

def requestFilingListPage(CIK, t):
    '''This function is an intermediary between finding the company CIK and pulling the 10-K document.
    It is called in the first line of get10kFilingInfo, not in the main function.
    
    In order to access any 10-K, we have to find the page where the 10-K's are listed. It takes the CIK
    number that getCIK() found, and it can also take a second argument that defines the type of document.
    Later on, we can start pulling 10-Q documents, etc.
    
    The function uses requests to go find the page on SEC Edgar. It then uses BeautifulSoup to turn this
    content into something the program can read through and parse.'''
    # define the endpoint to do filing searches.
    browse_edgar = r"https://www.sec.gov/cgi-bin/browse-edgar"

    # define the arguments of the request
    search_params = {
        'CIK':CIK,
        'Count':'100',
        'myowner':'include',
        'action':'getcompany',
        'type':t,
        'output':'atom',
    }

    # make the request
    response = requests.get(url=browse_edgar, params=search_params)
    soup = BeautifulSoup(response.content,'xml')
    #print("Response URL:\n"+response.url)
    
    return soup

def getFilingInfo(CIK, counter, t):
    '''This function uses the requestFilingListPage() function to find a list of filings published by the
    company. By default, it finds 10-k filings. However, this can be changed from the main function to find
    other types of documents, such as 10-Q reports, etc.

    It then searches through the soup to find all 'entries,' or recordings of financial reports. Each entry
    has a different accession number. So, the function loops through each entry to append each accession number
    to a list of acc numbers. The 0 element in this list is the most recently published report. The i argument
    finds the most recent report by default, but it can be changed in the main function.

    To get report data on previous reports as well as the most recent, we can loop through this function so that
    i increases by 1 each time.'''
    soup = requestFilingListPage(CIK, t)
    
    # find all the accession number entry tags
    entries = soup.find_all('entry')

    # initialize lists for storage
    masterListXML = []
    ACClist = []

    # loop through each entry
    for entry in entries:

        # grab the accession number to create a key value
        accession_num = entry.find('accession-number').text #don't know why, but SEC misspelled "number" here
        ACClist.append(accession_num)

        #Don't really need this dictionary, but it will likely be useful at some point.
        entry_dict = {}
        entry_dict[accession_num] = {}

        # store the file info
        entry_dict[accession_num]['filing_date'] = entry.find('filing-date').text
        entry_dict[accession_num]['filing_href'] = entry.find('filing-href').text
        entry_dict[accession_num]['filing_type'] = entry.find('filing-type').text

        # store in the master list
        masterListXML.append(entry_dict)

    # return accession number of the most recent 10-k filing
    ACC = ACClist[counter]
    filingDate = masterListXML[counter][ACC]['filing_date']
    
    #print("master_list_xml:\n\n"+str(master_list_xml))
    return ACC, filingDate

def getCompanyInfo(CIK, ticker):
    '''This function is not crucial to making the program work. However, it gives us helpful company info
    like business and mailing addresses, its Standard Industrial Classification (SIC) information, the
    company's official name, its CIK, and its fyscal year end. Yes, we already have the CIK, but its helpful
    to have all this info in one place. We can add more info as needed.'''
    soup = requestFilingListPage(CIK, t = '10-k')
    
    # find all the accession number entry tags
    info = soup.find('company-info')

    # loop through each entry
    for item in info:

        info_dict = {}

        # store the addresses info (mailing and business)
        info_dict['addresses'] = {}

        info_dict['addresses']['mailing'] = {}
        info_dict['addresses']['mailing']['city'] = info.find('city').text
        info_dict['addresses']['mailing']['state'] = info.find('state').text
        info_dict['addresses']['mailing']['street1'] = info.find('street1').text
        if info.find('street2') is not None:
            info_dict['addresses']['mailing']['street2'] = info.find('street2').text
        info_dict['addresses']['mailing']['zip'] = info.find('zip').text

        info_dict['addresses']['business'] = {}
        info_dict['addresses']['business']['city'] = info.find('city').text
        info_dict['addresses']['business']['phone'] = info.find('phone').text
        info_dict['addresses']['business']['state'] = info.find('state').text
        info_dict['addresses']['business']['street1'] = info.find('street1').text
        if info.find('street2') is not None:
            info_dict['addresses']['business']['street2'] = info.find('street2').text
        info_dict['addresses']['business']['zip'] = info.find('zip').text

        # store the company's Standard Industrial Classification (SIC) info
        info_dict['sicInfo'] = {}
        info_dict['sicInfo']['SIC'] = info.find('assigned-sic').text
        info_dict['sicInfo']['SIC Description'] = info.find('assigned-sic-desc').text
        info_dict['sicInfo']['SIC URL'] = info.find('assigned-sic-href').text

        # store name info
        info_dict['nameInfo'] = {}
        info_dict['nameInfo']['Name'] = info.find('conformed-name').text
        '''can add code to also append former names to the dictionary if needed'''
        
        # store other company info
        info_dict['CIK'] = info.find('cik').text
        info_dict['ticker'] = ticker
        info_dict['FYE'] = info.find('fiscal-year-end').text

    # return accession number of the most recent 10-k filing
    #print("\n\ninfo_dict:\n\n"+str(info_dict))
    companyName = info_dict['nameInfo']['Name']
    return info_dict, companyName

"""
Grab the Filing XML Summary:
"""
def get_filing_summary(companyInfoDict, ACC, counter, filingDate, t):
    '''Now that we have the accession number of the filing that we want, we can get the filing summary.
    This function builds a URL with the CIK number and the accession number. The resulting page is the
    summary of the financial report. To fully understand the way this function works, it's best to look
    at the 'File Path' URL that this function creates and prints out (the variable is xml_summary).
    The function outputs this xml_summary of the report so that the next functions can parse the summary,
    find and grab the financial statements, and then scrape them into an organized dictionary.'''
    # define the base url needed to create the file url and set xml_summary to None.
    base_url = r"https://www.sec.gov/"
    xml_summary = None

    # convert a normal url to a 10k document landing page url
    normal_url = base_url + "Archives/edgar/data/" + companyInfoDict['CIK'] + "/" + ACC + ".txt"
    DocLandPage_url = normal_url.replace('-','').replace('.txt','/index.json')

    # request the url and decode it.
    content = requests.get(DocLandPage_url).json()
    #print(content)

    if counter > 0:
        print('\n'+'-'*100)
        print("NOTICE:")
        print("There was an issue with the contents of the 10-K report. Pulling the info for the next most recent report.")
        print('-'*100 + '\n')
        
    for file in content['directory']['item']:
        
        # Grab the filing summary and create a new url leading to the file so we can download it.
        if file['name'] == 'FilingSummary.xml':

            xml_summary = base_url + content['directory']['name'] + "/" + file['name']
            
            print('-' * 100)
            print('File Name: ' + file['name'])
            print('File Date: ' + filingDate)
            print('File Path: ' + xml_summary)

    if xml_summary is None:
        '''sometimes if there is no Filing Summary in the filing, it's an amendment to the previous filing.
        If this is the case, increase the counter by 1 and call the getFilingInfo() and get_filing_summary()
        functions again.'''
        counter += 1
        ACC, filingDate = getFilingInfo(companyInfoDict['CIK'], counter, t)
        xml_summary = get_filing_summary(companyInfoDict, ACC, counter, filingDate, t)
        if type(xml_summary) is tuple:
            '''I don't know why this creates a tuple'''
            xml_summary = xml_summary[0]
    return xml_summary, counter

"""
Parse the Filing Summary:
"""
def parse_filing_summary(xml_summary, companyInfoDict, ACC, filingDate, counter):
    '''Now that we have the xml filing summary, this function parses it. The filing is divided up
    into sections called reports on the summary. Each report has a short name, long name, position,
    category, and url. To analyze financial statements, we are only looking for 4 of the reports in
    the 'Statements' category: the balance sheet, income statement, cash flows statement, and statement
    of owners equity.
    
    At the end of this function, we make sure that the 'Statements' category is in the summary. If not,
    the filing is probably an appended version. Companies are allowed to publish an appended version
    of their filing later in the year if they need to wait for some reason to publish everything in their
    initial filing. If this is the case, the function increases the 'counter' variable by 1, and it then
    calls the getFilingInfo(), get_filing_summary(), and parse_filing_summary() functions again to get the
    next most recent filing (yes, this is recursive – is there a better way?). Remember that in the getFilingInfo()
    function, the default is to get the most recently published filing. Now with the counter variable as
    the argument, we get the next most recent filing. When we start getting multiple filings, we need to
    take this into account.'''
    # If we need to download the reports, we can define a new url that represents the filing folder:
    filingFolderURL = xml_summary.replace('FilingSummary.xml', '')

    # request and parse the content
    content = requests.get(xml_summary).content
    soup = BeautifulSoup(content, 'lxml')

    # find the 'myreports' tag because this contains all the individual reports submitted.
    reports = soup.find('myreports')

    # I want a list to store all the individual components of the report, so create the master list.
    master_reports = []

    # loop through each report in the 'myreports' tag but avoid the last one as this will cause an error.
    for report in reports.find_all('report')[:-1]:

        # let's create a dictionary to store all the different parts we need.
        report_dict = {}
        report_dict['name_short'] = report.shortname.text
        report_dict['name_long'] = report.longname.text
        report_dict['position'] = report.position.text
        report_dict['category'] = report.menucategory.text
        report_dict['url'] = filingFolderURL + report.htmlfilename.text

        # append the dictionary to the master list.
        master_reports.append(report_dict)

        '''if report_dict['category'] == "Statements":
            print('-'*100)
            print("BASE URL:\n",base_url + report.htmlfilename.text)
            print("LONGNAME:\n",report.longname.text)
            print("SHORTNAME:\n",report.shortname.text)
            print("MENU CATEGORY:\n",report.menucategory.text)
            print("POSITION:\n",report.position.text)'''

        # print the info to the user.
        '''print('-'*100)
        print("BASE URL:\n",base_url + report.htmlfilename.text)
        #print("LONGNAME:\n",report.longname.text)
        print("SHORTNAME:\n",report.shortname.text)
        print("MENU CATEGORY:\n",report.menucategory.text)
        print("POSITION:\n",report.position.text)'''

    # if the 10-k for some reason doesn't have the statements (if it's a 10-K/A), use the next most recent one
    catList = []
    for x in master_reports:
        catList.append(x['category'])
        
    #think about adding a while iteration
    if 'Statements' not in catList:
        counter += 1
        ACC, filingDate = getFilingInfo(companyInfoDict['CIK'], counter)
        xml_summary = get_filing_summary(companyInfoDict['CIK'], ACC, counter, filingDate)
        master_reports, counter, ACC, filingDate, xml_summary = parse_filing_summary(xml_summary, companyInfoDict['CIK'], ACC, filingDate, counter)
        if counter > 5:
            print("None of the entity's 6 most recent 10-K reports have financial statements.")
            return
        
    return master_reports, counter, ACC, filingDate, xml_summary
    
"""
Grab the Financial Statements:
"""
def grab_financial_statements(master_reports):
    '''Now that we have a master_reports list with multiple dictionaries (report_dict), we can find the financial
    statements that we want and put their URL's into a list. First, we define the statements that we want from the
    master_reports list.

    It is very common for a company to list their financial statements in different orders, so it is very important
    to record the order in which the statement URL's are listed. For this reason, we create the reportOrder
    dictionary. This order is very important in the next functions.

    The function then returns the list of URL's for each statement, as well as the reportOrder dictionary.'''
    # create the list to hold the statement urls
    statements_url = []

    item1 = None
    item2 = None
    item3 = None
    item4 = None

    # define the statements we want to look for (defined as item1 - item4)
    # consider changing master_reports to all caps if errors keep popping up and to decrease lines of code
    for v in master_reports:
        # BALANCE SHEET
        if "Balance" in v['name_short']:
            item1 = r"" + v['name_short']
            break
        elif "BALANCE" in v['name_short']:
            item1 = r"" + v['name_short']
            break

    for v in master_reports:
        # INCOME STATEMENT
        if "Operations" in v['name_short']:
            item2 = r"" + v['name_short']
            break
        if "OPERATIONS" in v['name_short']:
            item2 = r"" + v['name_short']
            break
        elif "Income" in v['name_short']:
            item2 = r"" + v['name_short']
            break
        elif "INCOME" in v['name_short']:
            item2 = r"" + v['name_short']
            break

    for v in master_reports:
        # STATEMENT OF CASH FLOWS
        if "Cash" in v['name_short']:
            item3 = r"" + v['name_short']
            break
        elif "CASH" in v['name_short']:
            item3 = r"" + v['name_short']
            break

    for v in master_reports:
        # STATEMENT OF STOCKHOLDERS EQUITY
        if "Equity" in v['name_short']:
            item4 = r"" + v['name_short']
            break
        elif "EQUITY" in v['name_short']:
            item4 = r"" + v['name_short']
            break
    '''print("item1: ",item1)
    print("item2: ",item2)
    print("item3: ",item3)
    print("item4: ",item4)'''

    # store the statement names in a list
    report_list = [item1, item2, item3, item4]
    reportOrder = {}
    n = 0
    #create a list with all the keys below in the desired order instead of iteration for report order
    for reportDict in master_reports:

        # if the short name can be found in the report list.
        if reportDict['name_short'] in report_list:

            # record index position of financial statement
            if "Balance" in reportDict['name_short']:
                reportOrder['Balance Sheet'] = n
            elif "BALANCE" in reportDict['name_short']:
                reportOrder['Balance Sheet'] = n
            elif "Operations" in reportDict['name_short']:
                reportOrder['Income Statement'] = n
            elif "OPERATIONS" in reportDict['name_short']:
                reportOrder['Income Statement'] = n
            elif "Income" in reportDict['name_short']:
                reportOrder['Income Statement'] = n
            elif "INCOME" in reportDict['name_short']:
                reportOrder['Income Statement'] = n
            elif "Cash" in reportDict['name_short']:
                reportOrder['Statement of Cash Flows'] = n
            elif "CASH" in reportDict['name_short']:
                reportOrder['Statement of Cash Flows'] = n
            elif "Equity" in reportDict['name_short']:
                reportOrder["Statement of Stockholders' Equity"] = n
            elif "EQUITY" in reportDict['name_short']:
                reportOrder["Statement of Stockholders' Equity"] = n
            n += 1

            # print some info and store it in the statements url.
            print('-'*100)
            print(reportDict['name_short'])
            print(reportDict['url'])
                
            statements_url.append(reportDict['url'])
        
    #print("Report List:\n",report_list)
    #print("\nReport Order Dictionary:\n"+str(reportOrder))
    return statements_url, reportOrder

"""
Scrape the Financial Statements:
"""
def scrape_financial_statements(statements_url, reportOrder):
    '''This function loops through the list of URL's created by the grab_financial_statements() function.
    For each URL, it requests the contents and uses BeautifulSoup to read it in html format. It puts content
    from statement headers into the headers list, statement sections into the sections list, and statement
    data into the data list. These 3 lists are inside of the statement_data dictionary. There's a different
    statement_data dictionary for each statement that the function loops through.

    If you're trying to understand the reportOrderKeys, you'll probably want to print reportOrder and
    reportOrderKeys. Since companies have different names for the same statements, this solves that issue. It
    may also be helpful to print out the statement_data dictionary at the end of the function to see how the
    naming of different items in the dictionary works.

    The final lines of the function that are blocked out also help to understand the contents of the
    statements_data list. This is essentially the master list with all the data from each statement that we
    want.'''
    # let's assume we want all the statements in a single data set.
    statements_data = []
    reportOrderKeys = list(reportOrder.keys())
    n = 0
    sectionData = {}

    # loop through each statement url
    for statement in statements_url:

        # define a dictionary that will store the different parts of the statement.
        statement_data = {}
        statement_data['headers'] = []
        statement_data['sections'] = []
        statement_data['data'] = []

        # count the rows of data under each section
        sectionData[reportOrderKeys[n]] = {}
        dataRows = 0
        
        # request the statement file content
        content = requests.get(statement).content
        report_soup = BeautifulSoup(content, 'html')

        # find all the rows, figure out what type of row it is, parse the elements, and store in the statement file list.
        for index, row in enumerate(report_soup.table.find_all('tr')):
            
            # first let's get all the elements.
            cols = row.find_all('td')
            
            # if it's a regular row and not a section or a table header
            if (len(row.find_all('th')) == 0 and len(row.find_all('strong')) == 0): 
                reg_row = [ele.text.strip() for ele in cols]
                statement_data['data'].append(reg_row)
                dataRows += 1
                
            # if it's a regular row and a section but not a table header
            elif (len(row.find_all('th')) == 0 and len(row.find_all('strong')) != 0):
                sec_row = cols[0].text.strip()
                statement_data['sections'].append(sec_row)
                
                sectionData[reportOrderKeys[n]][sec_row] = dataRows
                dataRows = 0
                
            # finally if it's not any of those it must be a header
            elif (len(row.find_all('th')) != 0):            
                hed_row = [ele.text.strip() for ele in row.find_all('th')]
                statement_data['headers'].append(hed_row)
                
            else:            
                print('We encountered an error.')
                
        n += 1

        # append it to the master list.
        statements_data.append(statement_data)
    # some things to print to better understand this function:
    print("reportOrder Dictionary:\n"+str(reportOrder))
    print("reportOrderKeys List:\n"+str(reportOrderKeys))
    print("sectionData Dictionary:\n"+str(sectionData))
    '''
    print("-"*100)
    for i in range(len(statements_data)):
        print("STATEMENTS DATA SECTION " + str(i) + ":\n")
        print("HEADERS:\n"+str(statements_data[i]['headers']))
        print("SECTIONS:\n"+str(statements_data[i]['sections']))
        print("DATA:\n"+str(statements_data[i]['data']))
        print("\n")
    print("-"*100)'''
    return statements_data

def popFootnotes(df, n, result, contents):
    # Get rid of footnotes – empty columns and bottom rows
    '''PROBLEM: sometimes a financial statement has a footnote attached that doesn't
    translate to pandas.
    
    SOLUTION: remove the column with the footnote reference, and all of the associated
    footnote data (probably in the bottom three rows) from the dataframe. It is important to store that
    deleted data somewhere else so that we can output it into excel or something.

    METHOD: First put the footnote data into a dictionary (a list or tuple may actually be better).
    The next step is to replace the footnote symbol [n] with NaN. There will now be an entire column
    of NaN values that can be deleted. Then, delete the rows with the footnote data. After the dataframe
    is saved to an excel worksheet, we can now insert the footnote data that was previously saved to
    a dictionary into the bottom of the worksheet. This is an important last step, given that footnote
    data is sometimes more relevant than the actual financial statement.
    
    This function finds the column in which the string '[n]' exists. It then finds the rows where
    '[n]' exists. Then it creates a dictionary with the contents of the footnotes (this part probably
    needs some work in case there is more than 1 footnote). Finally, it replaces the string '[n]' with
    NaN.'''
    
    # get list of columns that contains the value
    seriesObj = result.any()
    columnNames = list(seriesObj[seriesObj == True].index)
    #print("seriesObj:",seriesObj)
    #print("columnNames:",columnNames)
    
    # Iterate over list of columns and fetch the rows indexes where value exists
    for col in columnNames:
        rows = list(result[col][result[col] == True].index)
        for row in rows:
            listOfPos.append((row, col))
            
    # make a dictionary with the contents of the footnotes
    for x in range(len(listOfPos)):
        footnoteDict[n] = listOfPos[x][0] #need to check to make sure this works when there are multiple footnotes

    # replace [n] footnote symbols in data with NaN
    symbol = str('[' + str(n) + ']')
    df.replace(symbol, np.nan, inplace=True)
            
    # listOfPos is a list of tuples indicating the positions of footnotes in the dataframe
    '''print("\nList of positions:\n",listOfPos)
    print("\nLength of list:",len(listOfPos))
    print("\nfootnote dict:\n",footnoteDict)
    print("\nLength of dict:",len(footnoteDict))'''

    # delete rows with the remaining footnote data
    if n == 1:
        print("\nThere were no footnotes in the statement.\nContinuing execution as normal.\n")
    elif n > 1:
        print("\n\nFOOTNOTE FOUND")
        print("NUMBER OF FOOTNOTES IN STATEMENT:",n-1)
        print("DELETING THE BOTTOM",(n-1)*3,"LINES OF THE STATEMENT AS FOLLOWS:\n")
        for i in range((n-1)*3):
            indexes = df.index.tolist()
            rowCount = df.shape[0]
            print("\nRow number " + str(rowCount) + ":\n" + str(df.iloc[-1]))
            df.drop(indexes[-1], inplace=True)
        print("Rows successfully deleted")

    return df, footnoteDict

"""
Convert the Data into a Data Frame, then convert to excel:
"""
def make_balSheet_df(statements_data, reportOrder):
    '''In this function, we use Pandas to turn the balance sheet data (stored in the statements_data list)
    into a dataframe. If a footnote is detected, the popFootnotes() function is called.'''
    print("-"*50)
    print("MAKE BALANCE SHEET:")
    print("-"*50)
    
    # Grab the proper components
    bal_headerLists = statements_data[reportOrder['Balance Sheet']]['headers']
    #print(bal_headerLists)
    bal_header = bal_headerLists.pop(0) #bal_headerLists is for some reason a list within a list. This removes the outside list.
    #print("Header:\n"+str(bal_header))
    title = bal_header[0]
    bal_header = bal_header[1:]
    bal_data = statements_data[reportOrder['Balance Sheet']]['data']

    # Put the data in a DataFrame
    balSheet_df = pd.DataFrame(bal_data)
    
    '''# Display
    print('-'*100)
    print('Before Reindexing')
    print('-'*100)
    print(balSheet_df.head())'''
    
    # Define the Index column, rename it, and we need to make sure to drop the old column once we reindex.
    balSheet_df.index = balSheet_df[0]
    balSheet_df.index.name = 'Category'
    balSheet_df = balSheet_df.drop(0, axis = 1)
    
    '''# Display
    print('-'*100)
    print('Before Regex')
    print('-'*100)
    print(balSheet_df.head())'''
    
    # Get rid of the '$', '(', ')', and convert the '' to NaNs.
    balSheet_df = balSheet_df.replace('[\$,)]','', regex=True )\
                         .replace( '[(]','-', regex=True)\
                         .replace( '', 'NaN', regex=True)

    # Check for footnotes. If there are footnotes, the popFootnotes() function is called.
    listOfPos = []
    footnoteDict = {}
    n = 0
    while True:
        n+=1
        '''This iteration finds the '[n]' footnote if it exists. If not, the iteration ends. If it does exist,
        the popFootnotes() function is called.'''
        # get bool dataframe with True positions where the given value exists
        result = balSheet_df.isin(['[' + str(n) + ']'])
        contents = result.any(axis=None) #The while True iteration should stop after it finds all the footnotes.
        #print("Contents " + str(n) + ":\n" + str(contents))
        #print("Result " + str(n) + ":\n" + str(result))
        if contents == False:
            break
        balSheet_df, footnoteDict = popFootnotes(balSheet_df, n, result, contents)

    '''# Display
    print('-'*100)
    print('Before type conversion & removal of empty columns')
    print('-'*100)
    print(balSheet_df)'''

    # Convert data type from string to float
    balSheet_df = balSheet_df.astype(float)

    # Delete columns that are only NaN, and change the column headers
    balSheet_df.dropna(axis = 'columns', how = 'all', inplace=True)
    balSheet_df.columns = bal_header

    #Need to find a way to put section headers back into the dataframe...not of paramount importance

    # Display
    print('-'*100)
    print('Final Product')
    print('-'*100)
    print(balSheet_df)                
    
    print("BALANCE SHEET DATAFRAME SUCCESSFUL")
    return balSheet_df, footnoteDict

def make_income_df(statements_data, reportOrder, footnoteDict):
    '''In this function, we use Pandas to turn the income statement data (stored in the statements_data list)
    into a dataframe. If a footnote is detected, the popFootnotes() function is called.'''
    print("-"*50)
    print("MAKE INCOME STATEMENT:")
    print("-"*50)
    # Grab the proper components
    income_header =  statements_data[reportOrder['Income Statement']]['headers'][1]
    income_data = statements_data[reportOrder['Income Statement']]['data']
    #print("income header:\n",income_header)
    #print("income data:\n",income_data)

    # Put the data in a DataFrame
    income_df = pd.DataFrame(income_data)
    
    '''# Display
    print('-'*100)
    print('Before Reindexing')
    print('-'*100)
    print(income_df.head())'''
    
    # Define the Index column, rename it, and we need to make sure to drop the old column once we reindex.
    income_df.index = income_df[0]
    income_df.index.name = 'Category'
    income_df = income_df.drop(0, axis = 1)
    
    '''# Display
    print('-'*100)
    print('Before Regex')
    print('-'*100)
    print(income_df.head())'''
    
    # Get rid of the '$', '(', ')', and convert the '' to NaNs.
    income_df = income_df.replace('[\$,)]','', regex=True )\
                         .replace( '[(]','-', regex=True)\
                         .replace( '', 'NaN', regex=True)

    # Check for footnotes. If there are footnotes, the popFootnotes() function is called.
    listOfPos = []
    n = 0
    while True:
        n+=1
        '''This iteration finds the '[n]' footnote if it exists. If not, the iteration ends. If it does exist,
        the popFootnotes() function is called.'''
        # get bool dataframe with True positions where the given value exists
        result = income_df.isin(['[' + str(n) + ']'])
        contents = result.any(axis=None) #The while True iteration should stop after it finds all the footnotes.
        #print("Contents " + str(n) + ":\n" + str(contents))
        #print("Result " + str(n) + ":\n" + str(result))
        if contents == False:
            break
        income_df, footnoteDict = popFootnotes(income_df, n, result, contents)
    
    '''# Display
    print('-'*100)
    print('Before type conversion')
    print('-'*100)
    print(income_df.head())'''
    
    # everything is a string, so let's convert all the data to a float.
    income_df = income_df.astype(float)

    # Change the column headers
    income_df.columns = income_header

    # Display
    print('-'*100)
    print('Final Product')
    print('-'*100)
    print(income_df)    
    
    """# convert pandas dataframe into excel worksheet
    wsIncome = wb.create_sheet("Statement of Net Income (Loss)")
    for r in dataframe_to_rows(income_df, index=True, header=True):
        wsIncome.append(r)"""
    print("INCOME STATEMENT DATAFRAME SUCCESSFUL")
    return income_df, footnoteDict

def make_cashFlow_df(statements_data, reportOrder, footnoteDict):
    '''In this function, we use Pandas to turn the cash flow statement data (stored in the statements_data list)
    into a dataframe. If a footnote is detected, the popFootnotes() function is called.'''
    print("-"*50)
    print("MAKE CASH FLOW STATEMENT:")
    print("-"*50)
    # Grab the proper components
    cashFlow_header =  statements_data[reportOrder['Statement of Cash Flows']]['headers'][1]
    cashFlow_data = statements_data[reportOrder['Statement of Cash Flows']]['data']
    #print("Cash flow header:\n",cashFlow_header)
    #print("Cash flow data:\n",cashFlow_data)
    
    # Put the data in a DataFrame
    cashFlow_df = pd.DataFrame(cashFlow_data)
    
    '''# Display
    print('-'*100)
    print('Before Reindexing')
    print('-'*100)
    print(cashFlow_df.head())'''
    
    # Define the Index column, rename it, and we need to make sure to drop the old column once we reindex.
    cashFlow_df.index = cashFlow_df[0]
    cashFlow_df.index.name = 'Category'
    cashFlow_df = cashFlow_df.drop(0, axis = 1)
    
    '''# Display
    print('-'*100)
    print('Before Regex')
    print('-'*100)
    print(cashFlow_df.head())'''
    
    # Get rid of the '$', '(', ')', and convert the '' to NaNs.
    cashFlow_df = cashFlow_df.replace('[\$,)]','', regex=True )\
                         .replace( '[(]','-', regex=True)\
                         .replace( '', 'NaN', regex=True)

    # Check for footnotes. If there are footnotes, the popFootnotes() function is called.
    listOfPos = []
    n = 0
    while True:
        n+=1
        '''This iteration finds the '[n]' footnote if it exists. If not, the iteration ends. If it does exist,
        the popFootnotes() function is called.'''
        # get bool dataframe with True positions where the given value exists
        result = cashFlow_df.isin(['[' + str(n) + ']'])
        contents = result.any(axis=None) #The while True iteration should stop after it finds all the footnotes.
        #print("Contents " + str(n) + ":\n" + str(contents))
        #print("Result " + str(n) + ":\n" + str(result))
        if contents == False:
            break
        cashFlow_df, footnoteDict = popFootnotes(cashFlow_df, n, result, contents)
    
    '''# Display
    print('-'*100)
    print('Before type conversion')
    print('-'*100)
    print(cashFlow_df.head())'''
    
    # everything is a string, so let's convert all the data to a float.
    cashFlow_df = cashFlow_df.astype(float)

    # Change the column headers
    cashFlow_df.columns = cashFlow_header

    # Display
    print('-'*50)
    print('Final Product')
    print('-'*50)
    print(cashFlow_df)    
    
    """# convert pandas dataframe into excel worksheet
    wsCashFlow = wb.create_sheet("Statement of Cash Flows")
    for r in dataframe_to_rows(cashFlow_df, index=True, header=True):
        wsCashFlow.append(r)"""
    print("STATEMENT OF CASH FLOWS DATAFRAME SUCCESSFUL")
    return cashFlow_df, footnoteDict

def make_equity_df(statements_data, reportOrder, footnoteDict):
    '''In this function, we use Pandas to turn the statement of owners equity data (stored in
    the statements_data list) into a dataframe. If a footnote is detected, the popFootnotes()
    function is called.'''
    print("-"*50)
    print("MAKE EQUITY STATEMENT:")
    print("-"*50)
    # Grab the proper components
    equity_header =  statements_data[reportOrder["Statement of Stockholders' Equity"]]['headers'][0] #this is a list within a list. Should change this so that the [0] is not needed
    equity_header.pop(0) #removes the first element from the list
    print("equity_header:", equity_header)
    equity_data = statements_data[reportOrder["Statement of Stockholders' Equity"]]['data']
    #print("Equity header: ",equity_header)
    #print("Equity_data: ",equity_data)
    
    # Put the data in a DataFrame
    equity_df = pd.DataFrame(equity_data)
    
    # Display
    print('-'*100)
    print('Before Reindexing')
    print('-'*100)
    print(equity_df.head())
    
    # Define the Index column, rename it, and we need to make sure to drop the old column once we reindex.
    equity_df.index = equity_df[0]
    equity_df.index.name = 'Category'
    equity_df = equity_df.drop(0, axis = 1)
    
    # Display
    print('-'*100)
    print('Before Regex')
    print('-'*100)
    print(equity_df.head())
    
    # Get rid of the '$', '(', ')', and convert the '' to NaNs.
    equity_df = equity_df.replace('[\$,)]','', regex=True )\
                         .replace( '[(]','-', regex=True)\
                         .replace( '', 'NaN', regex=True)

    # Check for footnotes. If there are footnotes, the popFootnotes() function is called.
    listOfPos = []
    n = 0
    while True:
        n+=1
        '''This iteration finds the '[n]' footnote if it exists. If not, the iteration ends. If it does exist,
        the popFootnotes() function is called.'''
        # get bool dataframe with True positions where the given value exists
        result = equity_df.isin(['[' + str(n) + ']'])
        contents = result.any(axis=None) #The while True iteration should stop after it finds all the footnotes.
        #print("Contents " + str(n) + ":\n" + str(contents))
        #print("Result " + str(n) + ":\n" + str(result))
        if contents == False:
            break
        equity_df, footnoteDict = popFootnotes(equity_df, n, result, contents)
    
    # Display
    print('-'*100)
    print('Before type conversion')
    print('-'*100)
    print(equity_df.head())
    
    # everything is a string, so let's convert all the data to a float.
    equity_df = equity_df.astype(float)

    # Change the column headers
    equity_df.columns = equity_header

    # Display
    print('-'*100)
    print('Final Product')
    print('-'*100)
    print(equity_df)

    
    """# convert pandas dataframe into excel worksheet
    wsEquity = wb.create_sheet("Statement of Stockholders' Equity")
    for r in dataframe_to_rows(equity_df, index=True, header=True):
       wsEquity.append(r)"""
    print("STATEMENT OF STOCKHOLDERS' EQUITY DATAFRAME SUCCESSFUL")
    return equity_df, footnoteDict

"""
Perform Financial Analyses
"""

def analysis(balSheet_df, income_df, cashFlow_df, companyInfoDict, dirPath, filingDate, ticker):
    '''This function finds accounts from the balance sheet, income statement, and cash flow statement
    dataframes in order to create analyses on profitability, liquidity, and solvency. It outputs these
    calculations in the form of a csv document.'''
    balSheetHeaders = list(balSheet_df.columns)
    balSheetRows = list(balSheet_df.index)

    incomeHeaders = list(income_df.columns)
    incomeRows = list(income_df.index)

    cashRows = list(cashFlow_df.index)

    '''Profitability Analysis'''
    profitability_df = pd.DataFrame(index=['Revenue','Gross Profit','Gross Margin','EBITDA','EBIT','Operating Margin',
                                           'NOPAT','Net Income','Profit Margin','Return on Assets','Return on Equity'], columns=incomeHeaders)
    # find sales revenue
    rev1 = None
    rev2 = None
    rev3 = None
    revKeys = ["Revenue","Sales","Net sales","Total net sales","Net revenues","Net revenue","Total revenues",
               "operating revenues"]
    for x in range(len(revKeys)):
        '''this loop finds the index position of the revenue account on the income statement
        and returns the revenue for the 3 years reported in order from most recent to least
        recent'''
        revName = [i for i in incomeRows if revKeys[x] in i]
        if len(revName) == 1:
            rev1 = income_df.loc[revName[0],incomeHeaders[0]]
            rev2 = income_df.loc[revName[0],incomeHeaders[1]]
            rev3 = income_df.loc[revName[0],incomeHeaders[2]]
        elif len(revName) > 1:
            revNameLoc = incomeRows.index(revName[0])
            rev1 = income_df.iloc[revNameLoc,0]
            rev2 = income_df.iloc[revNameLoc,1]
            rev3 = income_df.iloc[revNameLoc,2]
            break

    # find COGS (cost of goods sold)
    cogs1 = None
    cogs2 = None
    cogs3 = None
    cogsKeys = ["Cost of sales","Costs of sales","Cost of revenue","Cost of goods sold","COGS","Cost of products sold",
                "cost of revenues","Selling, general and administrative expense",
                "Cost of Goods and Services Sold","Cost of services"]
    for x in range(len(cogsKeys)):
        cogsName = [i for i in incomeRows if cogsKeys[x] in i]
        if len(cogsName) == 1:
            cogs1 = income_df.loc[cogsName[0],incomeHeaders[0]]
            cogs2 = income_df.loc[cogsName[0],incomeHeaders[1]]
            cogs3 = income_df.loc[cogsName[0],incomeHeaders[2]]
        elif len(cogsName) > 1:
            cogsNameLoc = incomeRows.index(cogsName[0])
            cogs1 = income_df.iloc[cogsNameLoc,0]
            cogs2 = income_df.iloc[cogsNameLoc,1]
            cogs3 = income_df.iloc[cogsNameLoc,2]
            break

    # find operating income
    ebit1 = None
    ebit2 = None
    ebit3 = None
    ebitKeys = ["Income from operations","Operating income","Operating (loss) income","Operating Income",
                "Earnings from operations","OPERATING INCOME","Operating Profit",
                "Income from continuing operations before provision/(benefit) for taxes on income"]
    for x in range(len(ebitKeys)):
        ebitName = [i for i in incomeRows if ebitKeys[x] in i]
        if len(ebitName) == 1:
            ebit1 = income_df.loc[ebitName[0],incomeHeaders[0]]
            ebit2 = income_df.loc[ebitName[0],incomeHeaders[1]]
            ebit3 = income_df.loc[ebitName[0],incomeHeaders[2]]
        elif len(ebitName) > 1:
            ebitNameLoc = incomeRows.index(ebitName[0])
            ebit1 = income_df.iloc[ebitNameLoc,0]
            ebit2 = income_df.iloc[ebitNameLoc,1]
            ebit3 = income_df.iloc[ebitNameLoc,2]
            break

    # find net income
    income1 = None
    income2 = None
    income3 = None
    incomeKeys = ["Net income","Net Income","Net (loss) income","Net earnings"]
    for x in range(len(incomeKeys)):
        incomeName = [i for i in incomeRows if incomeKeys[x] in i]
        if len(incomeName) == 1:
            income1 = income_df.loc[incomeName[0],incomeHeaders[0]]
            income2 = income_df.loc[incomeName[0],incomeHeaders[1]]
            income3 = income_df.loc[incomeName[0],incomeHeaders[2]]
        elif len(incomeName) > 1:
            incomeNameLoc = incomeRows.index(incomeName[0])
            income1 = income_df.iloc[incomeNameLoc,0]
            income2 = income_df.iloc[incomeNameLoc,1]
            income3 = income_df.iloc[incomeNameLoc,2]
            break

    # find deprecitation expense
    dep1 = None
    dep2 = None
    dep3 = None
    depKeys = ["Depreciation"]
    for x in range(len(depKeys)):
        depName = [i for i in cashRows if depKeys[x] in i]
        if len(depName) == 1:
            dep1 = cashFlow_df.loc[depName[0],incomeHeaders[0]]
            dep2 = cashFlow_df.loc[depName[0],incomeHeaders[1]]
            dep3 = cashFlow_df.loc[depName[0],incomeHeaders[2]]
        elif len(depName) > 1:
            depNameLoc = cashRows.index(depName[0])
            dep1 = cashFlow_df.iloc[depNameLoc,0]
            dep2 = cashFlow_df.iloc[depNameLoc,1]
            dep3 = cashFlow_df.iloc[depNameLoc,2]
            break

    # find amortization expense
    am1 = None
    am2 = None
    am3 = None
    amKeys = ["Amortization","amortization"]
    for x in range(len(amKeys)):
        if amKeys[x] in depName[0]:
            am1 = 0
            am2 = 0
            am3 = 0
            break
    if am1 is None:
        for x in range(len(amKeys)):
            amName = [i for i in cashRows if amKeys[x] in i]
            if len(amName) == 1:
                am1 = cashFlow_df.loc[amName[0],incomeHeaders[0]]
                am2 = cashFlow_df.loc[amName[0],incomeHeaders[1]]
                am3 = cashFlow_df.loc[amName[0],incomeHeaders[2]]
            elif len(amName) > 1:
                amNameLoc = cashRows.index(amName[0])
                am1 = cashFlow_df.iloc[amNameLoc,0]
                am2 = cashFlow_df.iloc[amNameLoc,1]
                am3 = cashFlow_df.iloc[amNameLoc,2]
                break
    
    # find interest expense
    intExp1 = None
    intExp2 = None
    intExp3 = None
    intKeys = ["Interest expense"]
    for x in range(len(intKeys)):
        intName = [i for i in incomeRows if intKeys[x] in i]
        if len(intName) == 1:
            intExp1 = income_df.loc[intName[0],incomeHeaders[0]]
            intExp2 = income_df.loc[intName[0],incomeHeaders[1]]
            intExp3 = income_df.loc[intName[0],incomeHeaders[2]]
        elif len(intName) > 1:
            intNameLoc = incomeRows.index(intName[0])
            intExp1 = income_df.iloc[intNameLoc,0]
            intExp2 = income_df.iloc[intNameLoc,1]
            intExp3 = income_df.iloc[intNameLoc,2]
            break
        
    # find tax expense
    taxExp1 = None
    taxExp2 = None
    taxExp3 = None
    taxKeys = ["Provision for income taxes","Provision for/(Benefit from) income taxes",
               "Provision for/(benefit from) income taxes","Provision for taxes",
               "Provision/(benefit) for taxes on income","Income tax (benefit) expense",
               "(Provision) benefit for income taxes","Income taxes","Income tax expense (benefit)"]
    for x in range(len(taxKeys)):
        taxName = [i for i in incomeRows if taxKeys[x] in i]
        if len(taxName) == 1:
            taxExp1 = income_df.loc[taxName[0],incomeHeaders[0]]
            taxExp2 = income_df.loc[taxName[0],incomeHeaders[1]]
            taxExp3 = income_df.loc[taxName[0],incomeHeaders[2]]
        elif len(taxName) > 1:
            taxNameLoc = incomeRows.index(taxName[0])
            taxExp1 = income_df.iloc[taxNameLoc,0]
            taxExp2 = income_df.iloc[taxNameLoc,1]
            taxExp3 = income_df.iloc[taxNameLoc,2]
            break

    # find total assets
    totAss1 = None
    totAss2 = None
    totAssKeys = ["Total assets","Total Assets","TOTAL ASSETS"]
    for x in range(len(totAssKeys)):
        if balSheetRows.count(totAssKeys[x]) == 1:
            totAss1 = balSheet_df.loc[totAssKeys[x],balSheetHeaders[0]]
            totAss2 = balSheet_df.loc[totAssKeys[x],balSheetHeaders[1]]
    if totAss1 is None:
        for x in range(len(totAssKeys)):
            totAssName = [i for i in balSheetRows if totAssKeys[x] in i]
            if len(totAssName) > 0:
                totAssLoc = balSheetRows.index(totAssName[0])
                totAss1 = balSheet_df.iloc[totAssLoc,0]
                totAss2 = balSheet_df.iloc[totAssLoc,1]

    # find total equity
    totEq1 = None
    totEq2 = None
    totEqKeys = ["Total stockholders' equity","Total stockholders’ equity","Total shareholders' equity",
                   "Total Stockholders' Equity","Total shareholders’ equity","Total equity","Total Equity",
                   "TOTAL EQUITY"]
    for x in range(len(totEqKeys)):
        if totEqKeys[x] in balSheetRows:
            totEq1 = balSheet_df.loc[totEqKeys[x],balSheetHeaders[0]]
            totEq2 = balSheet_df.loc[totEqKeys[x],balSheetHeaders[1]]
    if totEq1 is None:
        for x in range(len(totEqKeys)):
            totEqName = [i for i in balSheetRows if totEqKeys[x] in i]
            if len(totEqName) > 0:
                totEqLoc = balSheetRows.index(totEqName[0])
                totEq1 = balSheet_df.iloc[totEqLoc,0]
                totEq2 = balSheet_df.iloc[totEqLoc,1]

    # Sales Revenue
    profitability_df.loc['Revenue',incomeHeaders[0]] = rev1
    profitability_df.loc['Revenue',incomeHeaders[1]] = rev2
    profitability_df.loc['Revenue',incomeHeaders[2]] = rev3

    # Gross Profit = Net Sales - COGS
    gp1 = rev1 - cogs1
    profitability_df.loc['Gross Profit',incomeHeaders[0]] = gp1

    gp2 = rev2 - cogs2
    profitability_df.loc['Gross Profit',incomeHeaders[1]] = gp2

    gp3 = rev3 - cogs3
    profitability_df.loc['Gross Profit',incomeHeaders[2]] = gp3

    # Gross Margin = Gross Profit / Net Sales
    gm1 = gp1 / rev1
    profitability_df.loc['Gross Margin',incomeHeaders[0]] = gm1

    gm2 = gp2 / rev2
    profitability_df.loc['Gross Margin',incomeHeaders[1]] = gm2

    gm3 = gp3 / rev3
    profitability_df.loc['Gross Margin',incomeHeaders[2]] = gm3

    # EBITDA = Net Income + Interest + Taxes + Depreciation + Amortization = Operating Profit + Depreciation + Amortization
    print(ebit1)
    print(dep1)
    print(am1)
    ebitda1 = ebit1 + dep1 + am1
    profitability_df.loc['EBITDA',incomeHeaders[0]] = ebitda1

    ebitda2 = ebit2 + dep2 + am2
    profitability_df.loc['EBITDA',incomeHeaders[1]] = ebitda2

    ebitda3 = ebit3 + dep3 + am3
    profitability_df.loc['EBITDA',incomeHeaders[2]] = ebitda3

    # EBIT = Operating Income (Operating Revenue - Operating Expenses)
    profitability_df.loc['EBIT',incomeHeaders[0]] = ebit1
    profitability_df.loc['EBIT',incomeHeaders[1]] = ebit2
    profitability_df.loc['EBIT',incomeHeaders[2]] = ebit3
    
    #NOPAT = Net Operating Profit After Tax = EBIT - Tax Expense
    nopat1 = ebit1 - taxExp1
    profitability_df.loc['NOPAT',incomeHeaders[0]] = nopat1
    
    nopat2 = ebit2 - taxExp2
    profitability_df.loc['NOPAT',incomeHeaders[1]] = nopat2
    
    nopat3 = ebit3 - taxExp3
    profitability_df.loc['NOPAT',incomeHeaders[2]] = nopat3

    # Operating Margin = EBIT / Sales Revenue
    opM1 = ebit1 / rev1
    profitability_df.loc['Operating Margin',incomeHeaders[0]] = opM1

    opM2 = ebit2 / rev2
    profitability_df.loc['Operating Margin',incomeHeaders[1]] = opM2

    opM3 = ebit3 / rev3
    profitability_df.loc['Operating Margin',incomeHeaders[2]] = opM3

    # Net Income
    profitability_df.loc['Net Income',incomeHeaders[0]] = income1
    profitability_df.loc['Net Income',incomeHeaders[1]] = income2
    profitability_df.loc['Net Income',incomeHeaders[2]] = income3
    
    # Profit margin = net income / sales revenue
    pm1 = income1 / rev1
    profitability_df.loc['Profit Margin',incomeHeaders[0]] = pm1
    
    pm2 = income2 / rev2
    profitability_df.loc['Profit Margin',incomeHeaders[1]] = pm2
    
    pm3 = income3 / rev3
    profitability_df.loc['Profit Margin',incomeHeaders[2]] = pm3
    
    #average total assets = (Beginning Total Assets + Ending Total Assets) / 2
    #To get the average total assets of 2017 and 2018, we'll need to pull the balance sheets for both of those years.
    avgTotAssets1 = (totAss1 + totAss2) / 2

    # Return on Assets (ROA) = net income / average total assets
    roa1 = income1/avgTotAssets1
    profitability_df.loc['Return on Assets',incomeHeaders[0]] = roa1

    #average total equity = (begginning total equity + ending total equity) / 2
    #To get the average total stockholders' equity of 2017 and 2018, we'll need to pull the balance sheets for those years.
    avgTotEquity1 = (totEq1 + totEq2) / 2
    
    # Return on Equity = Net Income / average stockholders' equity
    roe1 = income1/avgTotEquity1
    profitability_df.loc['Return on Equity',incomeHeaders[0]] = roe1

    print("\n\nProfitability Dataframe:\n",profitability_df)
    
    '''Liquidity Analysis'''
    liquidity_df = pd.DataFrame(index=['Working Capital','Current Ratio','Quick Ratio','Cash Ratio'], columns=balSheetHeaders)

    # find Cash & equivalents, assign to variable
    ce1 = None
    ce2 = None
    ceKeys = ["Cash and cash equivalents","CASH","Cash","cash"]
    for x in range(len(ceKeys)):
        ceName = [i for i in balSheetRows if ceKeys[x] in i]
        if len(ceName) == 1:
            ce1 = balSheet_df.loc[ceName[0],balSheetHeaders[0]]
            ce2 = balSheet_df.loc[ceName[0],balSheetHeaders[1]]
        elif len(ceName) > 1:
            ceNameLoc = balSheetRows.index(ceName[0])
            ce1 = balSheet_df.iloc[ceNameLoc,0]
            ce2 = balSheet_df.iloc[ceNameLoc,1]
            break

    # find marketable securities, assign to variable
    ms1 = None
    ms2 = None
    msKeys = ["Marketable securities","Trading assets","Short-term investments"]
    for x in range(len(msKeys)):
        msName = [i for i in balSheetRows if msKeys[x] in i]
        if len(msName) == 1:
            ms1 = balSheet_df.loc[msName[0],balSheetHeaders[0]]
            ms2 = balSheet_df.loc[msName[0],balSheetHeaders[1]]
        elif len(msName) > 1:
            msNameLoc = balSheetRows.index(msName[0])
            ms1 = balSheet_df.iloc[msNameLoc,0]
            ms2 = balSheet_df.iloc[msNameLoc,1]
            break
    if ms1 is None:
        ms1 = 0
        ms2 = 0
                
    # find accounts receivable, assign to variable
    ar1 = None
    ar2 = None
    arKeys = ["Accounts receivable","receivables","accounts receivable","Accounts and notes receivable"]
    for x in range(len(arKeys)):
        arName = [i for i in balSheetRows if arKeys[x] in i]
        if len(arName) == 1:
            ar1 = balSheet_df.loc[arName[0],balSheetHeaders[0]]
            ar2 = balSheet_df.loc[arName[0],balSheetHeaders[1]]
        elif len(arName) > 1:
            arNameLoc = balSheetRows.index(arName[0])
            ar1 = balSheet_df.iloc[arNameLoc,0]
            ar2 = balSheet_df.iloc[arNameLoc,1]
            break

    # find total current assets, assign to variable
    ca1 = None
    ca2 = None
    caKeys = ["Total current assets","TOTAL CURRENT ASSETS","Assets, Current"]
    for x in range(len(caKeys)):
        caName = [i for i in balSheetRows if caKeys[x] in i]
        if len(caName) == 1:
            ca1 = balSheet_df.loc[caName[0],balSheetHeaders[0]]
            ca2 = balSheet_df.loc[caName[0],balSheetHeaders[0]]
        elif len(caName) > 1:
            caNameLoc = balSheetRows.index(caName[0])
            ca1 = balSheet_df.iloc[caNameLoc,0]
            ca2 = balSheet_df.iloc[caNameLoc,1]
            break
        
    # find total current liabilities, assign to variable
    clKeys = ["Total current liabilities","TOTAL CURRENT LIABILITIES","Liabilities, Current"]
    for x in range(len(clKeys)):
        clName = [i for i in balSheetRows if clKeys[x] in i]
        if len(clName) == 1:
            cl1 = balSheet_df.loc[clName[0],balSheetHeaders[0]]
            cl2 = balSheet_df.loc[clName[0],balSheetHeaders[0]]
        elif len(clName) > 1:
            clNameLoc = balSheetRows.index(clName[0])
            cl1 = balSheet_df.iloc[clNameLoc,0]
            cl2 = balSheet_df.iloc[clNameLoc,1]
            break
        
    # working capital = Current Assets - Current Liabilities
    wc1 = ca1 - cl1
    liquidity_df.loc['Working Capital', balSheetHeaders[0]] = wc1

    wc2 = ca2 - cl2
    liquidity_df.loc['Working Capital', balSheetHeaders[1]] = wc2

    # current ratio = Current Assets / Current Liabilities
    cr1 = ca1/cl1
    liquidity_df.loc['Current Ratio', balSheetHeaders[0]] = cr1

    cr2 = ca2/cl2
    liquidity_df.loc['Current Ratio', balSheetHeaders[1]] = cr2

    # quick ratio = (cash and equivalents + marketable securities + accounts receivable) / current liabilities
    # OR            (current assets - inventory - prepaid expenses) / current liabilities
    qr1 = (ce1 + ms1 + ar1) / cl1
    liquidity_df.loc['Quick Ratio', balSheetHeaders[0]] = qr1

    qr2 = (ce2 + ms2 + ar2) / cl2
    liquidity_df.loc['Quick Ratio', balSheetHeaders[1]] = qr2

    # cash ratio = cash and equivalents / current liabilities
    cashR1 = ce1 / cl1
    liquidity_df.loc['Cash Ratio', balSheetHeaders[0]] = cashR1

    cashR2 = ce2 / cl2
    liquidity_df.loc['Cash Ratio', balSheetHeaders[1]] = cashR1
    
    print("\n\nLiquidity Dataframe:\n",liquidity_df)


    '''Solvency Analysis'''
    solvency_df = pd.DataFrame(index=['Debt-to-Equity Ratio','Interest Coverage Ratio'], columns=balSheetHeaders)

    # find total liabilities, assign to variable
    tl1 = None
    tl2 = None
    tse1 = None
    tse2 = None
    tlKeys = ["Total liabilities","Total Liabilities"]
    for x in range(len(tlKeys)):
        tlName = [i for i in balSheetRows if tlKeys[x] in i]
        if len(tlName) == 1:
            print("\n\nUsing '" + tlName[0] + "' as the account for Total Liabilities.")
            tl1 = balSheet_df.loc[tlName[0],balSheetHeaders[0]]
            tl2 = balSheet_df.loc[tlName[0],balSheetHeaders[1]]
        elif len(tlName) > 1:
            tlNameLoc = balSheetRows.index(tlName[0])
            tl1 = balSheet_df.iloc[tlNameLoc,0]
            tl2 = balSheet_df.iloc[tlNameLoc,1]
            break
    if len(tlName) == 0:
        #print("More work not in vain")
        tseKeys = ["Total stockholders' equity","Total stockholders’ equity","Total shareholders' equity",
                   "Total Stockholders' Equity","Total shareholders’ equity","Total equity","Total Equity",
                   "TOTAL EQUITY"]
        tlseKeys = ["Total liabilities and stockholders' equity","Total liabilities and stockholders’ equity",
                   "Total liabilities and shareholders' equity","Total liabilities and shareholders’ equity",
                   "Total liabilities and equity","Total Liabilities and Equity","TOTAL LIABILITIES AND EQUITY",
                   "Total Liabilities and Stockholders' Equity",
                   "Stockholders' Equity, Including Portion Attributable to Noncontrolling Interest",
                    "Liabilities and Equity"]
        for x in range(len(tseKeys)):
            tseName = [i for i in balSheetRows if tseKeys[x] in i]
            for y in range(len(tlseKeys)):
                tlseName = [i for i in balSheetRows if tlseKeys[y] in i]
                if len(tseName) and len(tlseName) > 0:
                    if tseName[0] and tlseName[0] in balSheetRows:
                        tse1 = balSheet_df.loc[tseName[0],balSheetHeaders[0]]
                        tlse1 = balSheet_df.loc[tlseName[0],balSheetHeaders[0]]
                        tse2 = balSheet_df.loc[tseName[0],balSheetHeaders[1]]
                        tlse2 = balSheet_df.loc[tlseName[0],balSheetHeaders[1]]

                        tl1 = tlse1 - tse1
                        tl2 = tlse2 - tse2
                    else:
                        print("Couldn't find total liabilities")

    # find total stockholders' equity, assign to variable
    if tse1 is None:
        tseKeys = ["Total stockholders' equity","Total stockholders’ equity","Total shareholders' equity",
                   "Total Stockholders' Equity","Total shareholders’ equity","Total equity","Total Equity",
                   "TOTAL EQUITY"]
        for x in range(len(tseKeys)):
            tseName = [i for i in balSheetRows if tseKeys[x] in i]
            if len(tseName) == 1:
                tse1 = balSheet_df.loc[tseName[0],balSheetHeaders[0]]
                tse2 = balSheet_df.loc[tseName[0],balSheetHeaders[1]]
            elif len(tseName) > 1:
                tseNameLoc = balSheetRows.index(tseName[0])
                tse1 = balSheet_df.iloc[tseNameLoc,0]
                tse2 = balSheet_df.iloc[tseNameLoc,1]
                break
    
    # Debt-to-equity ratio = total liabilities / total stockholders' equity
    dte1 = tl1/tse1
    solvency_df.loc['Debt-to-Equity Ratio', balSheetHeaders[0]] = dte1
    
    dte2 = tl2/tse2
    solvency_df.loc['Debt-to-Equity Ratio', balSheetHeaders[1]] = dte2

    # Interest coverage ratio = ebit / interest expense
    intCovRat1 = ebit1 / intExp1
    solvency_df.loc['Interest Coverage Ratio', balSheetHeaders[0]] = intCovRat1

    intCovRat2 = ebit2 / intExp2
    solvency_df.loc['Interest Coverage Ratio', balSheetHeaders[1]] = intCovRat2

    print("\n\nSolvency Dataframe:\n",solvency_df)
    
    # DuPont Analysis

    # Append dataframes together into one
    analysis_df = profitability_df.append(liquidity_df)
    analysis_df = analysis_df.append(solvency_df)
    
    # Write dataframe to csv format
    fileName = (ticker + '_' + filingDate + '_Analysis' + '.csv')
    filePath = dirPath + '/' + fileName
    analysis_df.to_csv(filePath)


"""
Name and Main functions
"""
def main():
    ticker = str(input("Enter the ticker: ")).upper() #takes ticker as input, converts to upper case
    counter = 0 #this will be used in getFilingInfo(), get_filing_summary() and parse_filing_summary()
    t = '10-k' #we are looking for 10-k filings
    
    # use ticker to find its associated CIK
    CIK = getCIK(ticker)
    
    # get Accession number of the most recent 10k filing
    ACC, filingDate = getFilingInfo(CIK, counter, t)

    # Create a directory path for the company's data to be stored if it doesn't already exist.
    dirPath = (os.path.expanduser('~/Desktop/Python Finance/Output 10-K Analysis/' + ticker))
    if not os.path.exists(dirPath):
        os.makedirs(dirPath)

    # make a dictionary of the company's info
    companyInfoDict, companyName = getCompanyInfo(CIK, ticker)

    # get_filing_summary()
    xml_summary, counter = get_filing_summary(companyInfoDict, ACC, counter, filingDate, t)
    
    # parse_filing_summary()
    master_reports, counter, ACC, filingDate, xml_summary = parse_filing_summary(xml_summary, CIK, ACC, filingDate, counter)

    #grab_financial_statements()
    statements_url, reportOrder = grab_financial_statements(master_reports)

    #scrape_financial_statements()
    statements_data = scrape_financial_statements(statements_url, reportOrder)

    # make the pandas dataframes
    balSheet_df, footnoteDict = make_balSheet_df(statements_data, reportOrder)
    print("\n\n")
    income_df, footnoteDict = make_income_df(statements_data, reportOrder, footnoteDict)
    print("\n\n")
    cashFlow_df, footnoteDict = make_cashFlow_df(statements_data, reportOrder, footnoteDict)
    print("\n\n")
    #equity_df, footnoteDict = make_equity_df(statements_data, reportOrder, footnoteDict)
    print("\n\n")

    # use the financial statement dataframes to perform analysis on the company; save analysis to a csv file.
    analysis(balSheet_df, income_df, cashFlow_df, companyInfoDict, dirPath, filingDate, ticker)

    # Save the workbook.
    #wbName = (ticker + '_' + filingDate + '_10-K Financials' + '.xlsx')
    #wbPath = dirPath + '/' + wbName
    #wb.save(wbPath)

if __name__ == "__main__":
    main()
