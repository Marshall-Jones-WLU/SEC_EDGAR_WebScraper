'''
Author: Marshall Jones
Filename: finTrial[i] / SEC 10-k Analysis

Description:
This program takes input from the user for the ticker symbol of a company. It then finds that
company's SEC Edgar CIK number. It then finds the Accession number of the company's most recently
published 10-k filing (this can also be changed based on date of release and type of filing). From
there, the get_filing_summary() function returns the summary of the filing in xml format. The
parse_filing_summary() function parses the filing summary and creates a list of all the different
reports, or sections of the filing. The grab_financial_statements() function finds the financial
statements that we want from the list of sections of the filing, and then returns a smaller list
of just the financial statement data that we want (currently the Balance Sheet, Income Statement,
Statement of Cash Flows, and Statement of Stockholders' Equity, but more can be added).

The scrape_financial_statements() function takes the data from each financial statement and organizes
it into a structure with column headers, section headers (on each row), and their associated data.
Then, there is another function for each statement that uses the pandas library to put the data into
a data frame, which can be viewed in an excel file. A new worksheet is saved for each of these
financial statements. These functions are called make_balSheet_df(), make_income_df(), make_cashFlow_df(),
and make_equity_df()

TODO:
Analysis worksheet (performed with classes)

Formatting (performed with classes)
'''
# import libraries
import os
import csv
import re, requests
import pandas as pd
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

# Make an excel workbook to store the raw statements. It is saved at the end of the main() function
wb = Workbook()

"""
Use input ticker to find the CIK number, then the Accession number for the most
recent 10k report, as well as the company's info
"""
def getCIK(ticker):
    URL = 'https://www.sec.gov/cgi-bin/browse-edgar?CIK={}&Find=Search&owner=exclude&action=getcompany'
    CIK_RE = re.compile(r'.*CIK=(\d{10}).*')
    cik_dict = {}
    f = requests.get(URL.format(ticker), stream = True)
    results = CIK_RE.findall(f.text)
    
    if len(results):
        results[0] = int(re.sub('\.[0]*', '.', results[0]))
        cik_dict[str(ticker).upper()] = str(results[0])
    
    CIK = cik_dict[ticker]
    return CIK

def request10kListPage(CIK):
    # define the endpoint to do filing searches.
    browse_edgar = r"https://www.sec.gov/cgi-bin/browse-edgar"

    # define the arguments of the request
    search_params = {
        'CIK':CIK,
        'Count':'100',
        'myowner':'include',
        'action':'getcompany',
        'type':'10-k',
        'output':'atom',
    }

    # make the request
    response = requests.get(url=browse_edgar, params=search_params)
    soup = BeautifulSoup(response.content,'xml')
    #print("Response URL:\n"+response.url)
    
    return soup

def get10kFilingInfo(CIK, i=0):
    soup = request10kListPage(CIK)
    
    # find all the accession number entry tags
    entries = soup.find_all('entry')

    # initialize lists for storage
    masterListXML = []
    ACClist = []

    # loop through each entry
    for entry in entries:

        # grab the accession number to create a key value
        accession_num = entry.find('accession-nunber').text #don't know why, but SEC misspelled "number" here
        ACClist.append(accession_num)

        #Don't really need this dictionary, but it will likely be useful at some point.
        entry_dict = {}
        entry_dict[accession_num] = {}

        # store the file info
        entry_dict[accession_num]['filing_date'] = entry.find('filing-date').text
        entry_dict[accession_num]['filing_href'] = entry.find('filing-href').text
        entry_dict[accession_num]['filing_type'] = entry.find('filing-type').text

        # store in the master list
        masterListXML.append(entry_dict)

    # return accession number of the most recent 10-k filing
    ACC = ACClist[i]
    filingDate = masterListXML[i][ACC]['filing_date']
    
    #print("master_list_xml:\n\n"+str(master_list_xml))
    return ACC, filingDate

def getCompanyInfo(CIK, ticker):
    soup = request10kListPage(CIK)
    
    # find all the accession number entry tags
    info = soup.find('company-info')

    # loop through each entry
    for item in info:

        info_dict = {}

        # store the addresses info (mailing and business)
        info_dict['addresses'] = {}

        info_dict['addresses']['mailing'] = {}
        info_dict['addresses']['mailing']['city'] = info.find('city').text
        info_dict['addresses']['mailing']['state'] = info.find('state').text
        info_dict['addresses']['mailing']['street1'] = info.find('street1').text
        if info.find('street2') is not None:
            info_dict['addresses']['mailing']['street2'] = info.find('street2').text
        info_dict['addresses']['mailing']['zip'] = info.find('zip').text

        info_dict['addresses']['business'] = {}
        info_dict['addresses']['business']['city'] = info.find('city').text
        info_dict['addresses']['business']['phone'] = info.find('phone').text
        info_dict['addresses']['business']['state'] = info.find('state').text
        info_dict['addresses']['business']['street1'] = info.find('street1').text
        if info.find('street2') is not None:
            info_dict['addresses']['business']['street2'] = info.find('street2').text
        info_dict['addresses']['business']['zip'] = info.find('zip').text

        # store the company's Standard Industrial Classification (SIC) info
        info_dict['sicInfo'] = {}
        info_dict['sicInfo']['SIC'] = info.find('assigned-sic').text
        info_dict['sicInfo']['SIC Description'] = info.find('assigned-sic-desc').text
        info_dict['sicInfo']['SIC URL'] = info.find('assigned-sic-href').text

        # store name info
        info_dict['nameInfo'] = {}
        info_dict['nameInfo']['Name'] = info.find('conformed-name').text
        '''can add code to also append former names to the dictionary if needed'''
        
        # store other company info
        info_dict['CIK'] = info.find('cik').text
        info_dict['FYE'] = info.find('fiscal-year-end').text

    # return accession number of the most recent 10-k filing
    #print("\n\ninfo_dict:\n\n"+str(info_dict))
    companyName = info_dict['nameInfo']['Name']
    return info_dict, companyName

"""
Grab the Filing XML Summary:
"""
def get_filing_summary(CIK, ACC, counter, filingDate):
    # define the base url needed to create the file url.
    base_url = r"https://www.sec.gov/"

    # convert a normal url to a 10k document landing page url
    normal_url = base_url + "Archives/edgar/data/" + CIK + "/" + ACC + ".txt"
    DocLandPage_url = normal_url.replace('-','').replace('.txt','/index.json')
    #DocLandPage_url = r"https://www.sec.gov/Archives/edgar/data/1265107/000126510719000004/index.json"

    # request the url and decode it.
    content = requests.get(DocLandPage_url).json()

    if counter > 0:
        print('\n'+'-'*100)
        print("NOTICE:")
        print("There was an issue with the contents of the 10-K report. Now pulling the info for the next most recent report.")
        print('-'*100 + '\n')
        
    for file in content['directory']['item']:
        
        # Grab the filing summary and create a new url leading to the file so we can download it.
        if file['name'] == 'FilingSummary.xml':

            xml_summary = base_url + content['directory']['name'] + "/" + file['name']
            
            print('-' * 100)
            print('File Name: ' + file['name'])
            print('File Date: ' + filingDate)
            print('File Path: ' + xml_summary)

    return xml_summary

"""
Parse the Filing Summary:
"""
def parse_filing_summary(xml_summary, CIK, ACC, filingDate, counter):
    # define a new base url that represents the filing folder. This will come
    # in handy when we need to download the reports.
    base_url = xml_summary.replace('FilingSummary.xml', '')

    # request and parse the content
    content = requests.get(xml_summary).content
    soup = BeautifulSoup(content, 'lxml')

    # find the 'myreports' tag because this contains all the individual reports submitted.
    reports = soup.find('myreports')

    # I want a list to store all the individual components of the report, so create the master list.
    master_reports = []

    # loop through each report in the 'myreports' tag but avoid the last one as this will cause an error.
    for report in reports.find_all('report')[:-1]:

        # let's create a dictionary to store all the different parts we need.
        report_dict = {}
        report_dict['name_short'] = report.shortname.text
        report_dict['name_long'] = report.longname.text
        report_dict['position'] = report.position.text
        report_dict['category'] = report.menucategory.text
        report_dict['url'] = base_url + report.htmlfilename.text

        # append the dictionary to the master list.
        master_reports.append(report_dict)

        '''if report_dict['category'] == "Statements":
            print('-'*100)
            print("BASE URL:\n",base_url + report.htmlfilename.text)
            print("LONGNAME:\n",report.longname.text)
            print("SHORTNAME:\n",report.shortname.text)
            print("MENU CATEGORY:\n",report.menucategory.text)
            print("POSITION:\n",report.position.text)'''

        # print the info to the user.
        '''print('-'*100)
        print("BASE URL:\n",base_url + report.htmlfilename.text)
        #print("LONGNAME:\n",report.longname.text)
        print("SHORTNAME:\n",report.shortname.text)
        print("MENU CATEGORY:\n",report.menucategory.text)
        print("POSITION:\n",report.position.text)'''

    # if the 10-k for some reason doesn't have the statements (if it's a 10-K/A), use the next most recent one
    catList = []
    for x in master_reports:
        catList.append(x['category'])
        
    #think about adding a while iteration
    if 'Statements' not in catList:
        counter += 1
        ACC, filingDate = get10kFilingInfo(CIK, counter)
        xml_summary = get_filing_summary(CIK, ACC, counter, filingDate)
        master_reports, counter, ACC, filingDate, xml_summary = parse_filing_summary(xml_summary, CIK, ACC, filingDate, counter)
        if counter > 5:
            print("None of the entity's 6 most recent 10-K reports have financial statements.")
            return
        
    return master_reports, counter, ACC, filingDate, xml_summary
    
"""
Grab the Financial Statements:
"""
def grab_financial_statements(master_reports):
    # create the list to hold the statement urls
    statements_url = []

    item1 = None
    item2 = None
    item3 = None
    item4 = None

    # define the statements we want to look for (defined as item1 - item4)
    # consider changing master_reports to all caps if errors keep popping up, or to decrease lines of code
    for v in master_reports:
        # BALANCE SHEET
        if "Balance" in v['name_short']:
            item1 = r"" + v['name_short']
            break
        elif "BALANCE" in v['name_short']:
            item1 = r"" + v['name_short']
            break

    for v in master_reports:
        # INCOME STATEMENT
        if "Operations" in v['name_short']:
            item2 = r"" + v['name_short']
            break
        if "OPERATIONS" in v['name_short']:
            item2 = r"" + v['name_short']
            break
        elif "Income" in v['name_short']:
            item2 = r"" + v['name_short']
            break
        elif "INCOME" in v['name_short']:
            item2 = r"" + v['name_short']
            break

    for v in master_reports:
        # STATEMENT OF CASH FLOWS
        if "Cash" in v['name_short']:
            item3 = r"" + v['name_short']
            break
        elif "CASH" in v['name_short']:
            item3 = r"" + v['name_short']
            break

    for v in master_reports:
        # STATEMENT OF STOCKHOLDERS EQUITY
        if "Equity" in v['name_short']:
            item4 = r"" + v['name_short']
            break
        elif "EQUITY" in v['name_short']:
            item4 = r"" + v['name_short']
            break
    '''print("item1: ",item1)
    print("item2: ",item2)
    print("item3: ",item3)
    print("item4: ",item4)'''

    # store the statement names in a list
    report_list = [item1, item2, item3, item4]
    reportOrder = {}
    n = 0
    #create a list with all the keys below in the desired order instead of iteration for report order
    for reportDict in master_reports:

        # if the short name can be found in the report list.
        if reportDict['name_short'] in report_list:

            # record index position of financial statement
            if "Balance" in reportDict['name_short']:
                reportOrder['Balance Sheet'] = n
            elif "BALANCE" in reportDict['name_short']:
                reportOrder['Balance Sheet'] = n
            elif "Operations" in reportDict['name_short']:
                reportOrder['Income Statement'] = n
            elif "OPERATIONS" in reportDict['name_short']:
                reportOrder['Income Statement'] = n
            elif "Income" in reportDict['name_short']:
                reportOrder['Income Statement'] = n
            elif "INCOME" in reportDict['name_short']:
                reportOrder['Income Statement'] = n
            elif "Cash" in reportDict['name_short']:
                reportOrder['Statement of Cash Flows'] = n
            elif "CASH" in reportDict['name_short']:
                reportOrder['Statement of Cash Flows'] = n
            elif "Equity" in reportDict['name_short']:
                reportOrder["Statement of Stockholders' Equity"] = n
            elif "EQUITY" in reportDict['name_short']:
                reportOrder["Statement of Stockholders' Equity"] = n
            n += 1

            # print some info and store it in the statements url.
            print('-'*100)
            print(reportDict['name_short'])
            print(reportDict['url'])
                
            statements_url.append(reportDict['url'])
        
    #print("Report List:\n",report_list)
    #print("\nReport Order Dic:\n"+str(reportOrder))
    return statements_url, reportOrder

"""
Scrape the Financial Statements:
"""
def scrape_financial_statements(statements_url, reportOrder):
    # let's assume we want all the statements in a single data set.
    statements_data = []
    reportOrderKeys = list(reportOrder.keys())
    n = 0
    sectionData = {}

    # loop through each statement url
    for statement in statements_url:

        # define a dictionary that will store the different parts of the statement.
        statement_data = {}
        statement_data['headers'] = []
        statement_data['sections'] = []
        statement_data['data'] = []

        # count the rows of data under each section
        sectionData[reportOrderKeys[n]] = {}
        #print(reportOrderKeys[n])
        dataRows = 0
        
        # request the statement file content
        content = requests.get(statement).content
        report_soup = BeautifulSoup(content, 'html')

        # find all the rows, figure out what type of row it is, parse the elements, and store in the statement file list.
        for index, row in enumerate(report_soup.table.find_all('tr')):
            
            # first let's get all the elements.
            cols = row.find_all('td')
            
            # if it's a regular row and not a section or a table header
            if (len(row.find_all('th')) == 0 and len(row.find_all('strong')) == 0): 
                reg_row = [ele.text.strip() for ele in cols]
                statement_data['data'].append(reg_row)
                dataRows += 1
                
            # if it's a regular row and a section but not a table header
            elif (len(row.find_all('th')) == 0 and len(row.find_all('strong')) != 0):
                sec_row = cols[0].text.strip()
                statement_data['sections'].append(sec_row)
                
                sectionData[reportOrderKeys[n]][sec_row] = dataRows
                dataRows = 0
                
            # finally if it's not any of those it must be a header
            elif (len(row.find_all('th')) != 0):            
                hed_row = [ele.text.strip() for ele in row.find_all('th')]
                statement_data['headers'].append(hed_row)
                
            else:            
                print('We encountered an error.')
                
        n += 1

        # append it to the master list.
        statements_data.append(statement_data)

    #print("sectionData Dictionary:\n"+str(sectionData))
    '''print("-"*100)
    for i in range(len(statements_data)):
        print("STATEMENTS DATA SECTION " + str(i) + ":\n")
        print("HEADERS:\n"+str(statements_data[i]['headers']))
        print("SECTIONS:\n"+str(statements_data[i]['sections']))
        print("DATA:\n"+str(statements_data[i]['data']))
        print("\n")
    print("-"*100)'''
    return statements_data

"""
Convert the Data into a Data Frame, then convert to excel:
"""
def make_balSheet_df(statements_data, reportOrder):
    print("-"*50)
    print("MAKE BALANCE SHEET:")
    print("-"*50)
    
    # Grab the proper components
    bal_headerLists = statements_data[reportOrder['Balance Sheet']]['headers']
    #print(bal_headerLists)
    bal_header = bal_headerLists.pop(0) #bal_headerLists is for some reason a list within a list. This removes the outside list.
    #print("Header:\n"+str(bal_header))
    title = bal_header[0]
    #bal_header[0] = np.nan
    bal_header = bal_header[1:]
    bal_data = statements_data[reportOrder['Balance Sheet']]['data']

    # Put the data in a DataFrame
    balSheet_df = pd.DataFrame(bal_data)
    
    '''# Display
    print('-'*100)
    print('Before Reindexing')
    print('-'*100)
    print(balSheet_df.head())'''
    
    # Define the Index column, rename it, and we need to make sure to drop the old column once we reindex.
    balSheet_df.index = balSheet_df[0]
    balSheet_df.index.name = 'Category'
    balSheet_df = balSheet_df.drop(0, axis = 1)
    
    '''# Display
    print('-'*100)
    print('Before Regex')
    print('-'*100)
    print(balSheet_df.head())'''
    
    # Get rid of the '$', '(', ')', and convert the '' to NaNs.
    balSheet_df = balSheet_df.replace('[\$,)]','', regex=True )\
                         .replace( '[(]','-', regex=True)\
                         .replace( '', 'NaN', regex=True)

    # Get rid of footnotes – empty columns and bottom rows
    '''PROBLEM: sometimes a financial statement has a footnote attached that doesn't
    translate to pandas.
    
    SOLUTION: remove the column with the footnote reference, and all of the associated
    footnote data (probably in the bottom three rows) from the dataframe.

    METHOD: First put the footnote data into a dictionary (a list or tuple may actually be better).
    The next step is to replace the footnote symbol [n] with NaN. There will now be an entire column
    of NaN values that can be deleted. Then, delete the rows with the footnote data. After the dataframe
    is saved to an excel worksheet, we can now insert the footnote data that was previously saved to
    a dictionary into the bottom of the worksheet. This is an important last step, given that footnote
    data is sometimes more relevant than the actual financial statement.'''
    listOfPos = list()
    n = 0
    footnoteDict = {}
    while True: #DEAL WITH JPM 10-K FOOTNOTE
        '''This iteration first finds the '[n]' footnote if it exists. If not, the iteration ends. It then
        finds the column in which the string '[n]' exists. It then finds the rows where '[n]' exists. Then
        it creates a dictionary with the contents of the footnotes (this part probably needs some work in
        case there is more than 1 footnote). Finally, it replaces the string '[n]' with NaN.'''
        n+=1
        # get bool dataframe with True positions where the given value exists
        result = balSheet_df.isin(['[' + str(n) + ']'])
        
        cont = result.any(axis=None) #The while True iteration should stop after it finds all the footnotes. 
        #print("Cont " + str(n) + ":\n" + str(cont))
        #print("Result " + str(n) + ":\n" + str(result))
        if cont == False:
            break
        
        # get list of columns that contains the value
        seriesObj = result.any()
        columnNames = list(seriesObj[seriesObj == True].index)
        #print("seriesObj:",seriesObj)
        #print("columnNames:",columnNames)
        
        # Iterate over list of columns and fetch the rows indexes where value exists
        for col in columnNames:
            rows = list(result[col][result[col] == True].index)
            for row in rows:
                listOfPos.append((row, col))
                
        # make a dictionary with the contents of the footnotes
        for x in range(len(listOfPos)):
            footnoteDict[n] = listOfPos[x][0] #need to check to make sure this works when there are multiple footnotes

        # replace [n] footnote symbols in data with NaN
        symbol = str('[' + str(n) + ']')
        balSheet_df.replace(symbol, np.nan, inplace=True)
            
    # listOfPos is a list of tuples indicating the positions of footnotes in the dataframe
    '''print("\nList of positions:\n",listOfPos)
    print("\nLength of list:",len(listOfPos))
    print("\nfootnote dict:\n",footnoteDict)
    print("\nLength of dict:",len(footnoteDict))'''

    # delete rows with the remaining footnote data
    if n == 1:
        print("\nThere were no footnotes in the statement.\nContinuing execution as normal.\n")
    elif n > 1:
        print("\n\nFOOTNOTE FOUND")
        print("NUMBER OF FOOTNOTES IN STATEMENT:",n-1)
        print("DELETING THE BOTTOM",(n-1)*3,"LINES OF THE STATEMENT AS FOLLOWS:\n")
        for i in range((n-1)*3):
            indexes = balSheet_df.index.tolist()
            rowCount = balSheet_df.shape[0]
            print("\nRow number " + str(rowCount) + ":\n" + str(balSheet_df.iloc[-1]))
            balSheet_df.drop(indexes[-1], inplace=True)
        print("Rows successfully deleted")

    '''# Display
    print('-'*100)
    print('Before type conversion & removal of empty columns')
    print('-'*100)
    print(balSheet_df)'''

    # Convert data type from string to float
    balSheet_df = balSheet_df.astype(float)

    # Delete columns that are only NaN, and change the column headers
    balSheet_df.dropna(axis = 'columns', how = 'all', inplace=True)
    balSheet_df.columns = bal_header

    '''Need to find a way to put section headers back into the dataframe'''

    # Display
    print('-'*100)
    print('Final Product')
    print('-'*100)
    print(balSheet_df)

    '''EXCEL FORMATTING – There is currently no need to convert this dataframe into the csv format because
    we can use the pandas dataframes to make analyses'''
    
    # convert pandas dataframe into excel worksheet
    rowCount = balSheet_df.shape[0]
    colCount = balSheet_df.shape[1]

    wsBal = wb.create_sheet("Balance Sheet")
    for r in dataframe_to_rows(balSheet_df, index=True, header=True):
        wsBal.append(r)

    '''Figure out how to change number format to accounting format,
    preferably not currency'''
    
    # format column widths based on the longest value in each column
    '''This needs work. I'd like to set a maximum width that a column
    can have (probably about 75). If a cell has more characters than
    can fit, then increase height of that row and wrap text to fit.'''
    dims = {}
    for row in wsBal.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        wsBal.column_dimensions[col].width = value

    # insert row above dataframe, merge top columns, add title
    wsBal.insert_rows(1)
    titleCells = 'A1:' + chr(65+(colCount-1)) + str(1)
    wsBal.merge_cells(titleCells)
    cell = wsBal['A1']
    cell.value = title

    # if there is a footnote, add it to the bottom of the sheet
    if n > 1:
        footStart = str(5+rowCount)
        footEnd = chr(65+(colCount-1)) + str(int(footStart)+6)
        footnoteCells = 'A' + footStart + ':' + footEnd
        wsBal.merge_cells(footnoteCells)
        cell = wsBal['A'+footStart]
        cell.value = footnoteDict[1]
        cell.alignment = Alignment(wrapText=True)
        cell.font = Font(size=10, italic=True)

        # if there is more than one footnote, continue
        if n > 2:
            for i in range(n-2):
                footStart = str(int(footStart)+5)
                footEnd = chr(65+(colCount-1)) + str(int(footStart)+6)
                footnoteCells = 'A' + footStart + ':' + footEnd
                wsBal.merge_cells(footnoteCells)
                cell = wsBal['A'+footStart]
                cell.value = footnoteDict[i+1]
                cell.alignment = Alignment(wrapText=True)
                
    
    print("BALANCE SHEET DATAFRAME SUCCESSFUL")
    return balSheet_df

def make_income_df(statements_data, reportOrder):
    print("-"*50)
    print("MAKE INCOME STATEMENT:")
    print("-"*50)
    # Grab the proper components
    income_header =  statements_data[reportOrder['Income Statement']]['headers'][1]
    income_data = statements_data[reportOrder['Income Statement']]['data']
    #print("income header:\n",income_header)
    #print("income data:\n",income_data)

    # Put the data in a DataFrame
    income_df = pd.DataFrame(income_data)
    
    '''# Display
    print('-'*100)
    print('Before Reindexing')
    print('-'*100)
    print(income_df.head())'''
    
    # Define the Index column, rename it, and we need to make sure to drop the old column once we reindex.
    income_df.index = income_df[0]
    income_df.index.name = 'Category'
    income_df = income_df.drop(0, axis = 1)
    
    '''# Display
    print('-'*100)
    print('Before Regex')
    print('-'*100)
    print(income_df.head())'''
    
    # Get rid of the '$', '(', ')', and convert the '' to NaNs.
    income_df = income_df.replace('[\$,)]','', regex=True )\
                         .replace( '[(]','-', regex=True)\
                         .replace( '', 'NaN', regex=True)
    
    '''# Display
    print('-'*100)
    print('Before type conversion')
    print('-'*100)
    print(income_df.head())'''
    
    # everything is a string, so let's convert all the data to a float.
    income_df = income_df.astype(float)

    # Change the column headers
    income_df.columns = income_header

    # Display
    print('-'*100)
    print('Final Product')
    print('-'*100)
    print(income_df)    
    
    # convert pandas dataframe into excel worksheet
    wsIncome = wb.create_sheet("Statement of Net Income (Loss)")
    for r in dataframe_to_rows(income_df, index=True, header=True):
        wsIncome.append(r)
    print("INCOME STATEMENT DATAFRAME SUCCESSFUL")
    return income_df

def make_cashFlow_df(statements_data, reportOrder):
    print("-"*50)
    print("MAKE CASH FLOW STATEMENT:")
    print("-"*50)
    # Grab the proper components
    cashFlow_header =  statements_data[reportOrder['Statement of Cash Flows']]['headers'][1]
    cashFlow_data = statements_data[reportOrder['Statement of Cash Flows']]['data']
    #print("Cash flow header:\n",cashFlow_header)
    #print("Cash flow data:\n",cashFlow_data)
    
    # Put the data in a DataFrame
    cashFlow_df = pd.DataFrame(cashFlow_data)
    
    # Display
    print('-'*100)
    print('Before Reindexing')
    print('-'*100)
    print(cashFlow_df.head())
    
    # Define the Index column, rename it, and we need to make sure to drop the old column once we reindex.
    cashFlow_df.index = cashFlow_df[0]
    cashFlow_df.index.name = 'Category'
    cashFlow_df = cashFlow_df.drop(0, axis = 1)
    
    # Display
    print('-'*100)
    print('Before Regex')
    print('-'*100)
    print(cashFlow_df.head())
    
    # Get rid of the '$', '(', ')', and convert the '' to NaNs.
    cashFlow_df = cashFlow_df.replace('[\$,)]','', regex=True )\
                         .replace( '[(]','-', regex=True)\
                         .replace( '', 'NaN', regex=True)
    
    # Display
    print('-'*100)
    print('Before type conversion')
    print('-'*100)
    print(cashFlow_df.head())
    
    # everything is a string, so let's convert all the data to a float.
    cashFlow_df = cashFlow_df.astype(float)

    # Change the column headers
    cashFlow_df.columns = cashFlow_header

    # Display
    print('-'*100)
    print('Final Product')
    print('-'*100)
    print(cashFlow_df)    
    
    # convert pandas dataframe into excel worksheet
    wsCashFlow = wb.create_sheet("Statement of Cash Flows")
    for r in dataframe_to_rows(cashFlow_df, index=True, header=True):
        wsCashFlow.append(r)
    print("STATEMENT OF CASH FLOWS DATAFRAME SUCCESSFUL")
    return wsCashFlow

def make_equity_df(statements_data, reportOrder):
    print("-"*50)
    print("MAKE EQUITY STATEMENT:")
    print("-"*50)
    # Grab the proper components
    equity_header =  statements_data[reportOrder["Statement of Stockholders' Equity"]]['headers'][0] #this is a list within a list. Should change this so that the [0] is not needed
    equity_header.pop(0) #removes the first element from the list
    print("equity_header:", equity_header)
    equity_data = statements_data[reportOrder["Statement of Stockholders' Equity"]]['data']
    #print("Equity header: ",equity_header)
    #print("Equity_data: ",equity_data)
    
    # Put the data in a DataFrame
    equity_df = pd.DataFrame(equity_data)
    
    # Display
    print('-'*100)
    print('Before Reindexing')
    print('-'*100)
    print(equity_df.head())
    
    # Define the Index column, rename it, and we need to make sure to drop the old column once we reindex.
    equity_df.index = equity_df[0]
    equity_df.index.name = 'Category'
    equity_df = equity_df.drop(0, axis = 1)
    
    # Display
    print('-'*100)
    print('Before Regex')
    print('-'*100)
    print(equity_df.head())
    
    # Get rid of the '$', '(', ')', and convert the '' to NaNs.
    equity_df = equity_df.replace('[\$,)]','', regex=True )\
                         .replace( '[(]','-', regex=True)\
                         .replace( '', 'NaN', regex=True)
    
    # Display
    print('-'*100)
    print('Before type conversion')
    print('-'*100)
    print(equity_df.head())
    
    # everything is a string, so let's convert all the data to a float.
    equity_df = equity_df.astype(float)

    # Change the column headers
    equity_df.columns = equity_header

    # Display
    print('-'*100)
    print('Final Product')
    print('-'*100)
    print(equity_df)

    
    # convert pandas dataframe into excel worksheet
    wsEquity = wb.create_sheet("Statement of Stockholders' Equity")
    for r in dataframe_to_rows(equity_df, index=True, header=True):
       wsEquity.append(r)
    print("STATEMENT OF STOCKHOLDERS' EQUITY DATAFRAME SUCCESSFUL")
    return wsEquity

"""
Perform Financial Analyses
"""

def analysis(statements_data, reportOrder, balSheet_df, income_df, companyInfoDict, dirPath, filingDate, ticker):
    '''Need to do more work on this function to make it work for intc, ma, bac, cmcsa'''
    incomeHeaders = list(income_df.columns)
    incomeRows = list(income_df.index)
    
    balSheetHeaders = list(balSheet_df.columns)
    balSheetRows = list(balSheet_df.index)
    print(type(balSheet_df.index))
    print(type(balSheetRows))
    #print("balSheetRows:\n"+str(balSheetRows))

    '''Profitability Analysis'''
    profitability_df = pd.DataFrame(index=['EBITDA','EBIT','NOPAT','Net Income','Profit Margin','Return on Assets','Return on Equity'], columns=incomeHeaders)
    # find net income
    if "Net income" in income_df.index:
        income1 = income_df.loc["Net income",income_df.columns[0]]
        income2 = income_df.loc["Net income",income_df.columns[1]]
        income3 = income_df.loc["Net income",income_df.columns[2]]
    # find operating income
    if "Income from operations" in income_df.index:
        ebit1 = income_df.loc["Income from operations",income_df.columns[0]]
        ebit2 = income_df.loc["Income from operations",income_df.columns[1]]
        ebit3 = income_df.loc["Income from operations",income_df.columns[2]]
    # find deprecitation expense
    
    # find amortization expense
    # find interest expense
    # find tax expense
    if "Provision for income taxes" in income_df.index:
        taxExp1 = income_df.loc["Provision for income taxes",income_df.columns[0]]
        taxExp2 = income_df.loc["Provision for income taxes",income_df.columns[1]]
        taxExp3 = income_df.loc["Provision for income taxes",income_df.columns[2]]

    # find sales revenue
    if "Revenue" in income_df.index:
        rev1 = income_df.loc["Revenue",income_df.columns[0]]
        rev2 = income_df.loc["Revenue",income_df.columns[1]]
        rev3 = income_df.loc["Revenue",income_df.columns[2]]

    # EBITDA = Net Income + Interest + Taxes + Depreciation + Amortization = Operating Profit + Depreciation + Amortization

    # EBIT = Operating Income (Operating Revenue - Operating Expenses)
    profitability_df.loc['EBIT',incomeHeaders[0]] = ebit1
    profitability_df.loc['EBIT',incomeHeaders[1]] = ebit2
    profitability_df.loc['EBIT',incomeHeaders[2]] = ebit3
    
    #NOPAT = Net Operating Profit After Tax = EBIT - Tax Expense
    nopat1 = ebit1 - taxExp1
    profitability_df.loc['NOPAT',incomeHeaders[0]] = nopat1
    
    nopat2 = ebit2 - taxExp2
    profitability_df.loc['NOPAT',incomeHeaders[1]] = nopat2
    
    nopat3 = ebit3 - taxExp3
    profitability_df.loc['NOPAT',incomeHeaders[2]] = nopat3

    # Net Income
    profitability_df.loc['Net Income',incomeHeaders[0]] = income1
    profitability_df.loc['Net Income',incomeHeaders[1]] = income2
    profitability_df.loc['Net Income',incomeHeaders[2]] = income3
    
    # Profit margin = net income / sales revenue
    pm1 = income1 / rev1
    profitability_df.loc['Profit Margin',incomeHeaders[0]] = pm1
    
    pm2 = income2 / rev2
    profitability_df.loc['Profit Margin',incomeHeaders[1]] = pm2
    
    pm3 = income3 / rev3
    profitability_df.loc['Profit Margin',incomeHeaders[2]] = pm3
    
    #average total assets = (Beginning Total Assets + Ending Total Assets) / 2
    # find the average total assets of the two years reported on the balance sheet
    '''To get the average total assets of 2018 and 2017, we'll need to pull the balance sheet for those years.'''
    if "Total assets" in balSheet_df.index:
        avgTotAssets1 = (balSheet_df.loc['Total assets',balSheetHeaders[0]] + balSheet_df.loc['Total assets',balSheetHeaders[1]]) / 2
    # Return on Assets (ROA) = net income / average total assets
    roa1 = income1/avgTotAssets1
    profitability_df.loc['Return on Assets',incomeHeaders[0]] = roa1

    #find the average total stockholders' equity of the two years reported on the balance sheet
    '''To get the average total stockholders' equity of 2018 and 2017, we'll need to pull the balance sheet for those years.'''
    if "Total stockholders' equity" in balSheet_df.index:
        avgTotEquity1 = (balSheet_df.loc["Total stockholders' equity",balSheetHeaders[0]] + balSheet_df.loc["Total stockholders' equity",balSheetHeaders[1]]) / 2
    # Return on Equity = Net Income / average stockholders' equity
    roe1 = income1/avgTotEquity1
    profitability_df.loc['Return on Equity',incomeHeaders[0]] = roe1

    print("\n\nProfitability Dataframe:\n",profitability_df)
    
    '''Liquidity Analysis'''
    liquidity_df = pd.DataFrame(index=['Working Capital','Current Ratio','Quick Ratio','Cash Ratio'], columns=balSheetHeaders)

    # find Cash & equivalents
    if "Cash and cash equivalents" in balSheet_df.index:
        ce1 = balSheet_df.loc["Cash and cash equivalents",balSheetHeaders[0]]
        ce2 = balSheet_df.loc["Cash and cash equivalents",balSheetHeaders[1]]
    # find marketable securities
    if "Marketable securities" in balSheet_df.index:
        ms1 = balSheet_df.loc["Marketable securities",balSheetHeaders[0]]
        ms2 = balSheet_df.loc["Marketable securities",balSheetHeaders[1]]
    # find accounts receivable
    # first create a list of possible strings that will match the account name, then create a loop through the list,
    #then for each string in the list, use a list comprehension that checks if the string is anywhere in the row titles
    arKeys = ["Accounts receivable"]
    for x in range(len(arKeys)):
        arName = [i for i in balSheetRows if arKeys[x] in i]
        if len(arName) > 0:
            ar1 = balSheet_df.loc[arName[0],balSheetHeaders[0]]
            ar2 = balSheet_df.loc[arName[0],balSheetHeaders[1]]
            break
    # find total current assets, assign to variable
    caKeys = ["Total current assets","TOTAL CURRENT ASSETS","Assets, Current"]
    for x in range(len(caKeys)):
        caName = [i for i in balSheetRows if caKeys[x] in i]
        if len(caName) > 0:
            ca1 = balSheet_df.loc[caName[0],balSheetHeaders[0]]
            ca2 = balSheet_df.loc[caName[0],balSheetHeaders[0]]
            break
    '''if "Total current assets" in balSheet_df.index:
        ca1 = balSheet_df.loc['Total current assets',balSheetHeaders[0]]
        ca2 = balSheet_df.loc['Total current assets',balSheetHeaders[1]]
    elif "TOTAL CURRENT ASSETS" in balSheet_df.index:
        ca1 = balSheet_df.loc['TOTAL CURRENT ASSETS',balSheetHeaders[0]]
        ca2 = balSheet_df.loc['TOTAL CURRENT ASSETS',balSheetHeaders[1]]
    elif "Assets, Current" in balSheet_df.index:
        ca1 = balSheet_df.loc['Assets, Current',balSheetHeaders[0]]
        ca2 = balSheet_df.loc['Assets, Current',balSheetHeaders[1]]
    else:
        print("Can't find total current assets")'''
        
    # find total current liabilities, assign to variable
    if "Total current liabilities" in balSheet_df.index:
        cl1 = balSheet_df.loc['Total current liabilities',balSheetHeaders[0]]
        cl2 = balSheet_df.loc['Total current liabilities',balSheetHeaders[1]]
    elif "TOTAL CURRENT LIABILITIES" in balSheet_df.index:
        cl1 = balSheet_df.loc['TOTAL CURRENT LIABILITIES',balSheetHeaders[0]]
        cl2 = balSheet_df.loc['TOTAL CURRENT LIABILITIES',balSheetHeaders[1]]
    elif "Liabilities, Current" in balSheet_df.index:
        cl1 = balSheet_df.loc['Liabilities, Current',balSheetHeaders[0]]
        cl2 = balSheet_df.loc['Liabilities, Current',balSheetHeaders[1]]
    else:
        print("Can't find total current liabilities")
        
    # working capital = Current Assets - Current Liabilities
    wc1 = ca1 - cl1
    liquidity_df.loc['Working Capital', balSheetHeaders[0]] = wc1

    wc2 = ca2 - cl2
    liquidity_df.loc['Working Capital', balSheetHeaders[1]] = wc2

    # current ratio = Current Assets / Current Liabilities
    cr1 = ca1/cl1
    liquidity_df.loc['Current Ratio', balSheetHeaders[0]] = cr1

    cr2 = ca2/cl2
    liquidity_df.loc['Current Ratio', balSheetHeaders[1]] = cr2

    # quick ratio = (cash and equivalents + marketable securities + accounts receivable) / current liabilities
    # OR            (current assets - inventory - prepaid expenses) / current liabilities
    
    print("\n\nLiquidity Dataframe:\n",liquidity_df)


    '''Solvency Analysis'''
    solvency_df = pd.DataFrame(index=['Debt-to-Equity Ratio','Interest Coverage Ratio'], columns=balSheetHeaders)

    # find total liabilities, assign to variable
    if "Total liabilities" in balSheetRows:
        tl1 = balSheet_df.loc['Total liabilities',balSheetHeaders[0]]
        tl2 = balSheet_df.loc['Total liabilities',balSheetHeaders[1]]
    elif "Total Liabilities" in balSheetRows:
        tl1 = balSheet_df.loc['Total Liabilities',balSheetHeaders[0]]
        tl2 = balSheet_df.loc['Total Liabilities',balSheetHeaders[1]]
    elif "Liabilities" in balSheetRows:
        tl1 = balSheet_df.loc['Liabilities',balSheetHeaders[0]]
        tl2 = balSheet_df.loc['Liabilities',balSheetHeaders[1]]
    elif "Total stockholders' equity" and "Total liabilities and stockholders' equity" in balSheetRows:
        tl1 = balSheet_df.loc["Total liabilities and stockholders' equity",balSheetHeaders[0]] - balSheet_df.loc["Total stockholders' equity",balSheetHeaders[0]]
        tl2 = balSheet_df.loc["Total liabilities and stockholders' equity",balSheetHeaders[1]] - balSheet_df.loc["Total stockholders' equity",balSheetHeaders[1]]
    elif "Total stockholders’ equity" and "Total liabilities and stockholders’ equity" in balSheetRows:
        tl1 = balSheet_df.loc['Total liabilities and stockholders’ equity',balSheetHeaders[0]] - balSheet_df.loc['Total stockholders’ equity',balSheetHeaders[0]]
        tl2 = balSheet_df.loc['Total liabilities and stockholders’ equity',balSheetHeaders[1]] - balSheet_df.loc['Total stockholders’ equity',balSheetHeaders[1]]
    elif "Total shareholders' equity" and "Total liabilities and shareholders' equity" in balSheetRows:
        tl1 = balSheet_df.loc["Total liabilities and shareholders' equity",balSheetHeaders[0]] - balSheet_df.loc["Total shareholders' equity",balSheetHeaders[0]]
        tl2 = balSheet_df.loc["Total liabilities and shareholders' equity",balSheetHeaders[1]] - balSheet_df.loc["Total shareholders' equity",balSheetHeaders[1]]
    elif "Total shareholders’ equity" and "Total liabilities and shareholders’ equity" in balSheetRows:
        tl1 = balSheet_df.loc['Total liabilities and shareholders’ equity',balSheetHeaders[0]] - balSheet_df.loc['Total shareholders’ equity',balSheetHeaders[0]]
        tl2 = balSheet_df.loc['Total liabilities and shareholders’ equity',balSheetHeaders[1]] - balSheet_df.loc['Total shareholders’ equity',balSheetHeaders[1]]
    elif "Total equity" and "Total liabilities and equity" in balSheetRows:
        tl1 = balSheet_df.loc['Total liabilities and equity',balSheetHeaders[0]] - balSheet_df.loc['Total equity',balSheetHeaders[0]]
        tl2 = balSheet_df.loc['Total liabilities and equity',balSheetHeaders[1]] - balSheet_df.loc['Total equity',balSheetHeaders[1]]
    elif "Total equity" and "Total Liabilities and Equity" in balSheetRows:
        tl1 = balSheet_df.loc['Total Liabilities and Equity',balSheetHeaders[0]] - balSheet_df.loc['Total equity',balSheetHeaders[0]]
        tl2 = balSheet_df.loc['Total Liabilities and Equity',balSheetHeaders[1]] - balSheet_df.loc['Total equity',balSheetHeaders[1]]
    elif "TOTAL EQUITY" and "TOTAL LIABILITIES AND EQUITY" in balSheetRows:
        tl1 = balSheet_df.loc["TOTAL LIABILITIES AND EQUITY",balSheetHeaders[0]] - balSheet_df.loc["TOTAL EQUITY",balSheetHeaders[0]]
        tl2 = balSheet_df.loc["TOTAL LIABILITIES AND EQUITY",balSheetHeaders[1]] - balSheet_df.loc["TOTAL EQUITY",balSheetHeaders[1]]
    elif "Total stockholders' equity" and "Total Liabilities and Stockholders' Equity" in balSheetRows:
        tl1 = balSheet_df.loc["Total Liabilities and Stockholders' Equity",balSheetHeaders[0]] - balSheet_df.loc["Total stockholders' equity",balSheetHeaders[0]]
        tl2 = balSheet_df.loc["Total Liabilities and Stockholders' Equity",balSheetHeaders[1]] - balSheet_df.loc["Total stockholders' equity",balSheetHeaders[1]]
    else:
        print("Can't find total liabilities")

    # find total stockholders' equity, assign to variable
    if "Total stockholders' equity" in balSheetRows:
        tse1 = balSheet_df.loc["Total stockholders' equity",balSheetHeaders[0]]
        tse2 = balSheet_df.loc["Total stockholders' equity",balSheetHeaders[1]]
    elif "Total stockholders’ equity" in balSheetRows:
        tse1 = balSheet_df.loc["Total stockholders’ equity",balSheetHeaders[0]]
        tse2 = balSheet_df.loc["Total stockholders’ equity",balSheetHeaders[1]]
    elif "Total stockholders’ (deficit) equity" in balSheetRows:
        tse1 = balSheet_df.loc["Total stockholders’ (deficit) equity",balSheetHeaders[0]]
        tse2 = balSheet_df.loc["Total stockholders’ (deficit) equity",balSheetHeaders[1]]
    elif "Total shareholders' equity" in balSheetRows:
        tse1 = balSheet_df.loc["Total shareholders' equity",balSheetHeaders[0]]
        tse2 = balSheet_df.loc["Total shareholders' equity",balSheetHeaders[1]]
    elif "Total shareholders’ equity" in balSheetRows:
        tse1 = balSheet_df.loc["Total shareholders’ equity",balSheetHeaders[0]]
        tse2 = balSheet_df.loc["Total shareholders’ equity",balSheetHeaders[1]]
    elif "Total equity" in balSheetRows:
        tse1 = balSheet_df.loc["Total equity",balSheetHeaders[0]]
        tse2 = balSheet_df.loc["Total equity",balSheetHeaders[1]]
    elif "Total Equity" in balSheetRows:
        tse1 = balSheet_df.loc["Total Equity",balSheetHeaders[0]]
        tse2 = balSheet_df.loc["Total Equity",balSheetHeaders[1]]
    elif "TOTAL EQUITY" in balSheetRows:
        tse1 = balSheet_df.loc["TOTAL EQUITY",balSheetHeaders[0]]
        tse2 = balSheet_df.loc["TOTAL EQUITY",balSheetHeaders[1]]
    elif "Stockholders' Equity, Including Portion Attributable to Noncontrolling Interest" in balSheetRows:
        tse1 = balSheet_df.loc["Stockholders' Equity, Including Portion Attributable to Noncontrolling Interest",balSheetHeaders[0]]
        tse2 = balSheet_df.loc["Stockholders' Equity, Including Portion Attributable to Noncontrolling Interest",balSheetHeaders[1]]
    else:
        print("Can't find total stockholders' equity")
    
    # Debt-to-equity ratio = total liabilities / total stockholders' equity
    dte1 = tl1/tse1
    solvency_df.loc['Debt-to-Equity Ratio', balSheetHeaders[0]] = dte1
    
    dte2 = tl2/tse2
    solvency_df.loc['Debt-to-Equity Ratio', balSheetHeaders[1]] = dte2

    # Interest coverage ratio = (net income + interest expense + tax expense) / interest payments


    print("\n\nSolvency Dataframe:\n",solvency_df)
    
    # DuPont Analysis
    
    # Write dataframes to csv format    
    fileName = (ticker + '_' + filingDate + '_Liquidity Analysis' + '.csv')
    filePath = dirPath + '/' + fileName
    liquidity_df.to_csv(filePath)
    
    fileName = (ticker + '_' + filingDate + '_Solvency Analysis' + '.csv')
    filePath = dirPath + '/' +  fileName
    solvency_df.to_csv(filePath)
    
    
def makeAnalysisWS():
    pass
    ws = wb.active
    start = 1
    end = 5
    rows = ['A','B','C']
    # this loop merges the first 5 columns of the top 3 rows in the worksheet
    for i in range(len(rows)):
        c = chr(64+start)+str(i+1)+':'+chr(64+end)+str(i+1)
        ws.merge_cells(c)
    ws['A1'] = companyInfoDict['nameInfo']['Name']
    ws['A2'] = "Financial Statement Analysis"
    ws['A3'] = "From the "+filingDate+" 10-K Statement"

    ws['A5'] = "Liquidity Analysis:"
    liquidity_df.to_excel(writer, startrow=6)

    ws['A10'] = "Solvency Analysis:"
    solvency_df.to_excel(writer, startrow=11)
    writer.save()


"""
Name and Main functions
"""
def main():
    ticker = str(input("Enter the ticker: ")).upper()
    counter = 0
    
    # ask user for ticker, then find its associated CIK
    CIK = getCIK(ticker)
    
    # get Accession number of the most recent 10k report
    ACC, filingDate = get10kFilingInfo(CIK)

    # Create a directory path for the company's data to be stored if it doesn't already exist.
    dirPath = (os.path.expanduser('~/Desktop/Python Finance/Output 10-K Analysis/' + ticker))
    if not os.path.exists(dirPath):
        os.makedirs(dirPath)

    # make a dictionary of the company's info
    companyInfoDict, companyName = getCompanyInfo(CIK, ticker)

    # get_filing_summary()
    xml_summary = get_filing_summary(CIK,ACC,counter,filingDate)
    
    # parse_filing_summary()
    master_reports, counter, ACC, filingDate, xml_summary = parse_filing_summary(xml_summary, CIK, ACC, filingDate, counter)

    #grab_financial_statements()
    statements_url, reportOrder = grab_financial_statements(master_reports)

    #scrape_financial_statements()
    statements_data = scrape_financial_statements(statements_url, reportOrder)

    # make the pandas dataframes, and save them to a new worksheet in the excel workbook
    balSheet_df = make_balSheet_df(statements_data, reportOrder)
    print("\n\n")
    income_df = make_income_df(statements_data, reportOrder)
    print("\n\n")
    #wsCashFlow = make_cashFlow_df(statements_data, reportOrder)
    print("\n\n")
    #make_equity_df(statements_data, reportOrder)
    print("\n\n")
    analysis(statements_data, reportOrder, balSheet_df, income_df, companyInfoDict, dirPath, filingDate, ticker)
    #makeAnalysisWS(statements_data, reportOrder, balSheet_df, companyInfoDict, filingDate)

    # Save the workbook.
    wbName = (ticker + '_' + filingDate + '_10-K Financials' + '.xlsx')
    wbPath = dirPath + '/' + wbName
    wb.save(wbPath)

if __name__ == "__main__":
    main()
