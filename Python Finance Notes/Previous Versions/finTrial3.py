# import libraries
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook

# go ahead and make an excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Analysis"

# TODO: Short term goal is to be able to plug in the company ticker without manually entering
# the CIK or Accession number.
CIK = "1318605"
ACC_num = "0001564590-20-004475"

"""
Grab the Filing XML Summary:
"""

# define the base url needed to create the file url.
xml_base_url = r"https://www.sec.gov/"
base_url = r"https://www.sec.gov/Archives/edgar/data/"

# convert a normal url to a 10k document landing page url
normal_url = base_url + CIK + "/" + ACC_num + ".txt"
DocLandPage_url = normal_url.replace('-','').replace('.txt','/index.json')

# request the url and decode it.
content = requests.get(DocLandPage_url).json()

for file in content['directory']['item']:
        
    # Grab the filing summary and create a new url leading to the file so we can download it.
    if file['name'] == 'FilingSummary.xml':

        xml_summary = xml_base_url + content['directory']['name'] + "/" + file['name']
            
        print('-' * 100)
        print('File Name: ' + file['name'])
        print('File Path: ' + xml_summary)

"""
Parse the Filing Summary:
"""
# define a new base url that represents the filing folder. This will come
# in handy when we need to download the reports.
base_url = xml_summary.replace('FilingSummary.xml', '')
print("URL: ",base_url)

# request and parse the content
content = requests.get(xml_summary).content
print("Content: ",content)
soup = BeautifulSoup(content, 'lxml')
print("Soup: ",soup)

# find the 'myreports' tag because this contains all the individual reports submitted.
reports = soup.find('myreports')
print("Reports: ",reports)
# I want a list to store all the individual components of the report, so create the master list.
master_reports = []

# loop through each report in the 'myreports' tag but avoid the last one as this will cause an error.
for report in reports.find_all('report')[:-1]:

    # let's create a dictionary to store all the different parts we need.
    report_dict = {}
    report_dict['name_short'] = report.shortname.text
    report_dict['name_long'] = report.longname.text
    report_dict['position'] = report.position.text
    report_dict['category'] = report.menucategory.text
    report_dict['url'] = base_url + report.htmlfilename.text

    # append the dictionary to the master list.
    master_reports.append(report_dict)

    '''# print the info to the user.
    print('-'*100)
    print(base_url + report.htmlfilename.text)
    print(report.longname.text)
    print(report.shortname.text)
    print(report.menucategory.text)
    print(report.position.text)'''
    
"""
Grab the Financial Statements:
"""
# create the list to hold the statement urls
statements_url = []

for report_dict in master_reports:
        
    # define the statements we want to look for.
    item1 = r"Consolidated Balance Sheets"
    item2 = r"Consolidated Statements of Operations and Comprehensive Income (Loss)"
    item3 = r"Consolidated Statements of Cash Flows"
    item4 = r"Consolidated Statements of Stockholder's (Deficit) Equity"
        
    # store them in a list.
    report_list = [item1, item2, item3, item4]
        
    # if the short name can be found in the report list.
    if report_dict['name_short'] in report_list:
            
        # print some info and store it in the statements url.
        print('-'*100)
        print(report_dict['name_short'])
        print(report_dict['url'])
            
        statements_url.append(report_dict['url'])
        

"""
Scrape the Financial Statements:
"""
# let's assume we want all the statements in a single data set.
statements_data = []

# loop through each statement url
for statement in statements_url:

    # define a dictionary that will store the different parts of the statement.
    statement_data = {}
    statement_data['headers'] = []
    statement_data['sections'] = []
    statement_data['data'] = []
        
    # request the statement file content
    content = requests.get(statement).content
    report_soup = BeautifulSoup(content, 'html')

    # find all the rows, figure out what type of row it is, parse the elements, and store in the statement file list.
    for index, row in enumerate(report_soup.table.find_all('tr')):
            
        # first let's get all the elements.
        cols = row.find_all('td')
            
        # if it's a regular row and not a section or a table header
        if (len(row.find_all('th')) == 0 and len(row.find_all('strong')) == 0): 
            reg_row = [ele.text.strip() for ele in cols]
            statement_data['data'].append(reg_row)
                
        # if it's a regular row and a section but not a table header
        elif (len(row.find_all('th')) == 0 and len(row.find_all('strong')) != 0):
            sec_row = cols[0].text.strip()
            statement_data['sections'].append(sec_row)
                
        # finally if it's not any of those it must be a header
        elif (len(row.find_all('th')) != 0):            
            hed_row = [ele.text.strip() for ele in row.find_all('th')]
            statement_data['headers'].append(hed_row)
                
        else:            
            print('We encountered an error.')

    # append it to the master list.
    statements_data.append(statement_data)


"""
Convert the Data into a Data Frame:
"""
def make_balSheet_df():
    # Grab the proper components
    bal_header =  statements_data[1]['headers'][1]
    bal_data = statements_data[1]['data']

    # Put the data in a DataFrame
    balSheet_df = pd.DataFrame(bal_data)
    '''
    # Display
    print('-'*100)
    print('Before Reindexing')
    print('-'*100)
    print(balSheet_df.head())
    '''
    # Define the Index column, rename it, and we need to make sure to drop the old column once we reindex.
    balSheet_df.index = balSheet_df[0]
    balSheet_df.index.name = 'Category'
    balSheet_df = balSheet_df.drop(0, axis = 1)
    '''
    # Display
    print('-'*100)
    print('Before Regex')
    print('-'*100)
    print(balSheet_df.head())
    '''
    # Get rid of the '$', '(', ')', and convert the '' to NaNs.
    balSheet_df = balSheet_df.replace('[\$,)]','', regex=True )\
                         .replace( '[(]','-', regex=True)\
                         .replace( '', 'NaN', regex=True)
    '''
    # Display
    print('-'*100)
    print('Before type conversion')
    print('-'*100)
    print(balSheet_df.head())
    '''
    # everything is a string, so let's convert all the data to a float.
    balSheet_df = balSheet_df.astype(float)

    # Change the column headers
    balSheet_df.columns = bal_header

    # Display
    print('-'*100)
    print('Final Product')
    print('-'*100)

    print(balSheet_df)
    # balSheet_df.to_csv('income_state.csv')
    #wsBal = wb.create_sheet("Balance Sheet")
    #return wsBal

def make_income_df():
    # Grab the proper components
    income_header =  statements_data[1]['headers'][1]
    income_data = statements_data[1]['data']

    # Put the data in a DataFrame
    income_df = pd.DataFrame(income_data)
    '''
    # Display
    print('-'*100)
    print('Before Reindexing')
    print('-'*100)
    print(income_df.head())
    '''
    # Define the Index column, rename it, and we need to make sure to drop the old column once we reindex.
    income_df.index = income_df[0]
    income_df.index.name = 'Category'
    income_df = income_df.drop(0, axis = 1)
    '''
    # Display
    print('-'*100)
    print('Before Regex')
    print('-'*100)
    print(income_df.head())
    '''
    # Get rid of the '$', '(', ')', and convert the '' to NaNs.
    income_df = income_df.replace('[\$,)]','', regex=True )\
                         .replace( '[(]','-', regex=True)\
                         .replace( '', 'NaN', regex=True)
    '''
    # Display
    print('-'*100)
    print('Before type conversion')
    print('-'*100)
    print(income_df.head())
    '''
    # everything is a string, so let's convert all the data to a float.
    income_df = income_df.astype(float)

    # Change the column headers
    income_df.columns = income_header

    # Display
    print('-'*100)
    print('Final Product')
    print('-'*100)

    print(income_df)
    #income_df.to_excel('income_state.xlsx')
    #wsIncome = wb.create_sheet("Statement of Net Income (Loss)")
    #return wsIncome

def make_cashFlow_df():
    # Grab the proper components
    cashFlow_header =  statements_data[1]['headers'][1]
    cashFlow_data = statements_data[1]['data']

    # Put the data in a DataFrame
    cashFlow_df = pd.DataFrame(cashFlow_data)
    '''
    # Display
    print('-'*100)
    print('Before Reindexing')
    print('-'*100)
    print(cashFlow_df.head())
    '''
    # Define the Index column, rename it, and we need to make sure to drop the old column once we reindex.
    cashFlow_df.index = cashFlow_df[0]
    cashFlow_df.index.name = 'Category'
    cashFlow_df = cashFlow_df.drop(0, axis = 1)
    '''
    # Display
    print('-'*100)
    print('Before Regex')
    print('-'*100)
    print(cashFlow_df.head())
    '''
    # Get rid of the '$', '(', ')', and convert the '' to NaNs.
    cashFlow_df = cashFlow_df.replace('[\$,)]','', regex=True )\
                         .replace( '[(]','-', regex=True)\
                         .replace( '', 'NaN', regex=True)
    '''
    # Display
    print('-'*100)
    print('Before type conversion')
    print('-'*100)
    print(cashFlow_df.head())
    '''
    # everything is a string, so let's convert all the data to a float.
    cashFlow_df = cashFlow_df.astype(float)

    # Change the column headers
    cashFlow_df.columns = cashFlow_header

    # Display
    print('-'*100)
    print('Final Product')
    print('-'*100)

    print(cashFlow_df)
    # cashFlow_df.to_excel('income_state.xlsx')
    #wsCashFlow = wb.create_sheet("Statement of Cash Flows")
    #return wsCashFlow

def make_equity_df():
    # Grab the proper components
    equity_header =  statements_data[1]['headers'][1]
    equity_data = statements_data[1]['data']

    # Put the data in a DataFrame
    equity_df = pd.DataFrame(equity_data)
    '''
    # Display
    print('-'*100)
    print('Before Reindexing')
    print('-'*100)
    print(equity_df.head())
    '''
    # Define the Index column, rename it, and we need to make sure to drop the old column once we reindex.
    equity_df.index = equity_df[0]
    equity_df.index.name = 'Category'
    equity_df = equity_df.drop(0, axis = 1)
    '''
    # Display
    print('-'*100)
    print('Before Regex')
    print('-'*100)
    print(equity_df.head())
    '''
    # Get rid of the '$', '(', ')', and convert the '' to NaNs.
    equity_df = equity_df.replace('[\$,)]','', regex=True )\
                         .replace( '[(]','-', regex=True)\
                         .replace( '', 'NaN', regex=True)
    '''
    # Display
    print('-'*100)
    print('Before type conversion')
    print('-'*100)
    print(equity_df.head())
    '''
    # everything is a string, so let's convert all the data to a float.
    equity_df = equity_df.astype(float)

    # Change the column headers
    equity_df.columns = equity_header

    # Display
    print('-'*100)
    print('Final Product')
    print('-'*100)

    print(equity_df)
    #income_df.to_excel('income_state.xlsx')
    #wsEquity = wb.create_sheet("Statement of Stockholders' Equity")
    #return wsEquity

"""
Format Excel Document
"""



"""
Perform Financial Analyses
"""



"""
Name and Main functions
"""
def main():
    
    make_balSheet_df()
    make_income_df()
    make_cashFlow_df()
    make_equity_df()
    wb.save("finPyXLPractice.xlsx")

if __name__ == "__main__":
    main()
